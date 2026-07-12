from __future__ import annotations

import copy
import csv
import hashlib
import json
import re
import unicodedata
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

# ---------------------------------------------------------------------------
# 语义库加载
# ---------------------------------------------------------------------------

MODULE_DIR = Path(__file__).resolve().parent
DEFAULT_LIBRARY_CANDIDATES = (
    MODULE_DIR.parent / "source" / "semantic_library.json",
    MODULE_DIR / "semantic_library.json",
    MODULE_DIR.parent / "example_source" / "semantic" / "semantic_library.json",
)

DOMAIN_ORDER = ("地层", "岩性", "构造", "地下水", "不良地质", "围岩等级", "工程地质指标")
CATEGORY_LABELS = {
    "strata": "地层",
    "lithology": "岩性",
    "structure": "构造",
    "groundwater": "地下水",
    "adverse_geo": "不良地质",
    "rock_class": "围岩等级",
    "engineering_indicator": "工程地质指标",
}
DOMAIN_ALIASES = {
    **CATEGORY_LABELS,
    "formation": "地层",
    "stratigraphy": "地层",
    "rock": "岩性",
    "litho": "岩性",
    "fault": "构造",
    "water": "地下水",
    "hazard": "不良地质",
    "surrounding_rock": "围岩等级",
    "indicator": "工程地质指标",
    **{name: name for name in DOMAIN_ORDER},
}

ENCODING_SYSTEMS = {
    "PROJECT": "工具内部稳定编码",
    "SYMBOL": "常见地质符号",
    "CANONICAL": "规范名称",
    "ORIGINAL": "原始值",
}


def _fallback_library() -> dict[str, Any]:
    """外部语义库缺失时的最小兜底，保证程序仍能启动。"""
    return {
        "metadata": {
            "name": "最小兜底语义库",
            "version": "fallback-1.0",
            "code_notice": "请将 source/semantic_library.json 放入项目后使用完整语义库。",
        },
        "domains": {
            "地层": {
                "第四系": {
                    "aliases": ["第四纪", "Quaternary", "Q"],
                    "codes": {"PROJECT": "STR-QUATERNARY", "SYMBOL": "Q"},
                    "description": "第四纪地层。",
                }
            },
            "岩性": {
                "砂岩": {
                    "aliases": ["Sandstone"],
                    "codes": {"PROJECT": "LIT-SED-SANDSTONE"},
                    "description": "碎屑沉积岩。",
                },
                "石灰岩": {
                    "aliases": ["灰岩", "Limestone"],
                    "codes": {"PROJECT": "LIT-SED-LIMESTONE"},
                    "description": "碳酸盐岩。",
                },
            },
            "构造": {
                "断层": {
                    "aliases": ["Fault", "断裂"],
                    "codes": {"PROJECT": "GEO-STR-FAULT"},
                    "description": "断裂构造。",
                }
            },
            "地下水": {
                "裂隙水": {
                    "aliases": ["Fissure Water"],
                    "codes": {"PROJECT": "GW-FISSURE"},
                    "description": "赋存于裂隙中的地下水。",
                }
            },
            "不良地质": {
                "滑坡": {
                    "aliases": ["Landslide"],
                    "codes": {"PROJECT": "AG-LANDSLIDE"},
                    "description": "斜坡灾害。",
                }
            },
            "围岩等级": {
                "III级围岩": {
                    "aliases": ["Ⅲ级围岩", "Class III"],
                    "codes": {"PROJECT": "RC-III"},
                    "description": "三级围岩。",
                }
            },
            "工程地质指标": {
                "单轴抗压强度": {
                    "aliases": ["UCS", "抗压强度"],
                    "codes": {"PROJECT": "EI-UCS"},
                    "unit": "MPa",
                    "description": "单轴抗压强度。",
                }
            },
        },
    }


def load_semantic_library(library_path: str | Path | None = None) -> dict[str, Any]:
    candidates = [Path(library_path)] if library_path else list(DEFAULT_LIBRARY_CANDIDATES)
    for candidate in candidates:
        if not candidate.exists():
            continue
        try:
            data = json.loads(candidate.read_text(encoding="utf-8-sig"))
            domains = data.get("domains") if isinstance(data, dict) else None
            if isinstance(domains, dict) and domains:
                data.setdefault("metadata", {})["loaded_from"] = str(candidate.resolve())
                return data
        except (OSError, UnicodeDecodeError, json.JSONDecodeError):
            continue
    return _fallback_library()


_LIBRARY = load_semantic_library()
LIBRARY_METADATA: dict[str, Any] = _LIBRARY.get("metadata", {})
ALL_DICTIONARIES: dict[str, dict[str, dict[str, Any]]] = {
    domain: copy.deepcopy(_LIBRARY.get("domains", {}).get(domain, {})) for domain in DOMAIN_ORDER
}

# 兼容原主程序中的常量导入
STRATA_DICTIONARY = ALL_DICTIONARIES["地层"]
LITHOLOGY_DICTIONARY = ALL_DICTIONARIES["岩性"]
STRUCTURE_DICTIONARY = ALL_DICTIONARIES["构造"]
GROUNDWATER_DICTIONARY = ALL_DICTIONARIES["地下水"]
ADVERSE_GEO_DICTIONARY = ALL_DICTIONARIES["不良地质"]
ROCK_CLASS_DICTIONARY = ALL_DICTIONARIES["围岩等级"]
ENGINEERING_INDICATORS_DICTIONARY = ALL_DICTIONARIES["工程地质指标"]


# ---------------------------------------------------------------------------
# 基础工具
# ---------------------------------------------------------------------------

_NULL_VALUES = {"", "none", "null", "nan", "-", "--", "/", "—", "无", "未填"}


def _normalize_token(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"[\s_\-—–·,，;；:：/\\()\[\]{}<>《》]+", "", text)
    return text


def _canonical_domain(domain: str | None) -> str | None:
    if domain is None:
        return None
    value = str(domain).strip()
    return DOMAIN_ALIASES.get(value, DOMAIN_ALIASES.get(value.lower()))


def _as_alias_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    text = str(value).strip()
    if not text:
        return []
    return [item.strip() for item in re.split(r"[|,，;；\n]+", text) if item.strip()]


def _entry_codes(entry: dict[str, Any]) -> dict[str, str]:
    codes = entry.get("codes") if isinstance(entry.get("codes"), dict) else {}
    result = {str(k).upper(): str(v).strip() for k, v in codes.items() if str(v).strip()}
    # 兼容旧版 gb_code 字段，但不再把它自动宣称为国家标准
    if entry.get("gb_code") and "PROJECT" not in result:
        result["PROJECT"] = str(entry["gb_code"]).strip()
    return result


def _merge_dictionary_sets(
    base: dict[str, dict[str, dict[str, Any]]],
    additions: dict[str, dict[str, dict[str, Any]]] | None,
) -> dict[str, dict[str, dict[str, Any]]]:
    merged = {domain: copy.deepcopy(entries) for domain, entries in base.items()}
    if not additions:
        return merged
    for raw_domain, entries in additions.items():
        domain = _canonical_domain(raw_domain)
        if not domain or not isinstance(entries, dict):
            continue
        merged.setdefault(domain, {})
        for canonical, entry in entries.items():
            if not canonical or not isinstance(entry, dict):
                continue
            current = copy.deepcopy(merged[domain].get(canonical, {}))
            aliases = list(dict.fromkeys(_as_alias_list(current.get("aliases")) + _as_alias_list(entry.get("aliases"))))
            current.update(copy.deepcopy(entry))
            current["aliases"] = aliases
            old_codes = _entry_codes(merged[domain].get(canonical, {}))
            new_codes = _entry_codes(entry)
            if old_codes or new_codes:
                current["codes"] = {**old_codes, **new_codes}
            merged[domain][canonical] = current
    return merged


def semantic_library_stats(
    dictionaries: dict[str, dict[str, dict[str, Any]]] | None = None,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    dictionaries = dictionaries or ALL_DICTIONARIES
    per_domain: dict[str, Any] = {}
    total_terms = total_aliases = total_codes = 0
    for domain in DOMAIN_ORDER:
        entries = dictionaries.get(domain, {})
        aliases = sum(len(_as_alias_list(e.get("aliases"))) for e in entries.values())
        codes = sum(len(_entry_codes(e)) for e in entries.values())
        per_domain[domain] = {"terms": len(entries), "aliases": aliases, "codes": codes}
        total_terms += len(entries)
        total_aliases += aliases
        total_codes += codes
    return {
        "terms": total_terms,
        "aliases": total_aliases,
        "codes": total_codes,
        "domains": per_domain,
        "metadata": copy.deepcopy(metadata if metadata is not None else LIBRARY_METADATA),
    }


# ---------------------------------------------------------------------------
# 术语归一与编码映射
# ---------------------------------------------------------------------------


def _build_term_index(entries: dict[str, dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for canonical, entry in entries.items():
        tokens: list[tuple[str, str, str]] = [(canonical, "规范名", "CANONICAL")]
        tokens += [(a, "别名", "ALIAS") for a in _as_alias_list(entry.get("aliases"))]
        tokens += [(v, f"{k}编码", k) for k, v in _entry_codes(entry).items()]
        for token, matched_by, system in tokens:
            key = _normalize_token(token)
            if not key:
                continue
            candidate = {
                "canonical": canonical,
                "entry": entry,
                "token": str(token),
                "matched_by": matched_by,
                "system": system,
            }
            if all(existing["canonical"] != canonical or existing["matched_by"] != matched_by for existing in index[key]):
                index[key].append(candidate)
    return index


def normalize_name(
    raw: str,
    domain: str,
    custom_dict: dict[str, dict[str, Any]] | None = None,
    strategy: str = "规范优先",
    dictionaries: dict[str, dict[str, dict[str, Any]]] | None = None,
) -> dict[str, Any]:
    """将原始术语归一为规范名称，并返回候选、编码和关联关系。"""
    canonical_domain = _canonical_domain(domain)
    text = str(raw or "").strip()
    if not canonical_domain:
        return {
            "canonical": text,
            "matched": "语义域无效",
            "confidence": 0.0,
            "source": "original",
            "domain": domain,
            "candidates": [],
            "codes": {},
            "relations": [],
        }

    base_sets = dictionaries or ALL_DICTIONARIES
    entries = copy.deepcopy(base_sets.get(canonical_domain, {}))
    if custom_dict:
        entries = _merge_dictionary_sets({canonical_domain: entries}, {canonical_domain: custom_dict})[canonical_domain]
    index = _build_term_index(entries)
    token = _normalize_token(text)

    candidates = list(index.get(token, []))
    composite_matches: list[dict[str, Any]] = []
    match_kind = ""
    confidence = 0.0
    if candidates:
        # 精确规范名优先，其次编码，再其次别名。
        priority = {"规范名": 0}
        candidates.sort(
            key=lambda x: (
                priority.get(x["matched_by"], 1 if "编码" in x["matched_by"] else 2),
                x["canonical"],
            )
        )
        match_kind = candidates[0]["matched_by"] + "匹配"
        confidence = (
            1.0
            if candidates[0]["matched_by"] == "规范名"
            else 0.97
            if "编码" in candidates[0]["matched_by"]
            else 0.95
        )
    elif strategy != "保留原始" and len(token) >= 2 and canonical_domain != "围岩等级":
        # 复合描述允许同时识别多个术语，例如“砂岩夹泥岩”。
        # 围岩等级必须精确匹配，避免“VII级围岩”被错误识别为“I级围岩”。
        contained: list[dict[str, Any]] = []
        for key, options in index.items():
            if len(key) < 2:
                continue
            start_pos = token.find(key)
            if start_pos >= 0:
                for option in options:
                    option2 = dict(option)
                    option2["match_length"] = len(key)
                    option2["match_start"] = start_pos
                    option2["match_end"] = start_pos + len(key)
                    contained.append(option2)
        if contained:
            # 先保留长词，避免“石英砂岩”同时被拆成“砂岩”。
            contained.sort(key=lambda x: (-x["match_length"], x["match_start"], x["canonical"]))
            selected_ranges: list[tuple[int, int]] = []
            selected_by_name: dict[str, dict[str, Any]] = {}
            for item in contained:
                span = (item["match_start"], item["match_end"])
                if any(span[0] >= old[0] and span[1] <= old[1] for old in selected_ranges):
                    continue
                selected_ranges.append(span)
                selected_by_name.setdefault(item["canonical"], item)
            composite_matches = sorted(
                selected_by_name.values(),
                key=lambda x: (x["match_start"], -x["match_length"], x["canonical"]),
            )
            if strategy == "最长匹配":
                composite_matches = composite_matches[:1]
            match_kind = "最长子串匹配" if strategy == "最长匹配" else "复合描述匹配"
            confidence = 0.72

    if composite_matches:
        matched_names = [item["canonical"] for item in composite_matches]
        matched_entries = [entries.get(name, {}) for name in matched_names]
        aggregate_codes: dict[str, str] = {}
        for system in {key for entry in matched_entries for key in _entry_codes(entry)}:
            values = list(dict.fromkeys(_entry_codes(entry).get(system, "") for entry in matched_entries))
            aggregate_codes[system] = "|".join(value for value in values if value)
        relations: list[dict[str, Any]] = []
        for entry in matched_entries:
            if isinstance(entry.get("relations"), list):
                relations.extend(copy.deepcopy(entry["relations"]))
        return {
            "canonical": "；".join(matched_names),
            "matched": match_kind,
            "confidence": confidence,
            "source": "semantic-library",
            "domain": canonical_domain,
            "candidates": [],
            "matched_terms": matched_names,
            "codes": aggregate_codes,
            "relations": relations,
            "category": "；".join(dict.fromkeys(str(e.get("category", "")) for e in matched_entries if e.get("category"))),
            "subcategory": "；".join(dict.fromkeys(str(e.get("subcategory", "")) for e in matched_entries if e.get("subcategory"))),
            "unit": "；".join(dict.fromkeys(str(e.get("unit", "")) for e in matched_entries if e.get("unit"))),
            "description": "；".join(str(e.get("description", "")) for e in matched_entries if e.get("description")),
        }

    unique_candidates = list(dict.fromkeys(c["canonical"] for c in candidates))
    if not unique_candidates:
        return {
            "canonical": text,
            "matched": "未匹配",
            "confidence": 0.0,
            "source": "original",
            "domain": canonical_domain,
            "candidates": [],
            "codes": {},
            "relations": [],
        }

    if len(unique_candidates) > 1 and strategy == "保留原始":
        return {
            "canonical": text,
            "matched": "多候选-保留原始",
            "confidence": 0.0,
            "source": "original",
            "domain": canonical_domain,
            "candidates": unique_candidates,
            "codes": {},
            "relations": [],
        }

    selected = candidates[0]
    canonical = selected["canonical"]
    entry = entries.get(canonical, {})
    source = "custom-or-external" if custom_dict and canonical in custom_dict else "semantic-library"
    return {
        "canonical": canonical,
        "matched": match_kind,
        "confidence": confidence if len(unique_candidates) == 1 else max(0.55, confidence - 0.15),
        "source": source,
        "domain": canonical_domain,
        "candidates": unique_candidates,
        "codes": _entry_codes(entry),
        "relations": copy.deepcopy(entry.get("relations", [])) if isinstance(entry.get("relations"), list) else [],
        "category": entry.get("category", ""),
        "subcategory": entry.get("subcategory", ""),
        "unit": entry.get("unit", ""),
        "description": entry.get("description", ""),
    }


def map_encoding(
    value: str,
    from_system: str,
    to_system: str,
    domain: str = "岩性",
    dictionaries: dict[str, dict[str, dict[str, Any]]] | None = None,
) -> dict[str, Any]:
    """在原始值、规范名、项目码和地质符号之间进行实际映射。"""
    canonical_domain = _canonical_domain(domain) or domain
    dictionaries = dictionaries or ALL_DICTIONARIES
    entries = dictionaries.get(canonical_domain, {})
    src = str(from_system or "ORIGINAL").upper()
    dst = str(to_system or "PROJECT").upper()

    canonical = ""
    confidence = 0.0
    if src == "CANONICAL" and value in entries:
        canonical, confidence = value, 1.0
    elif src in {"PROJECT", "SYMBOL"}:
        token = _normalize_token(value)
        for name, entry in entries.items():
            if _normalize_token(_entry_codes(entry).get(src, "")) == token and token:
                canonical, confidence = name, 1.0
                break
    else:
        result = normalize_name(value, canonical_domain, dictionaries=dictionaries)
        canonical, confidence = result["canonical"], float(result["confidence"])

    entry = entries.get(canonical, {})
    if dst == "CANONICAL":
        mapped = canonical
    elif dst == "ORIGINAL":
        mapped = value
    else:
        mapped = _entry_codes(entry).get(dst, "")

    if canonical and mapped:
        return {
            "mapped_value": mapped,
            "canonical_name": canonical,
            "confidence": confidence,
            "route": f"{value} [{src}] → {canonical} → {mapped} [{dst}]",
            "from_system": src,
            "to_system": dst,
        }
    return {
        "mapped_value": value,
        "canonical_name": canonical or value,
        "confidence": confidence,
        "route": f"{value} [{src}] → 未找到 {dst} 映射",
        "from_system": src,
        "to_system": dst,
    }


def build_semantic_dictionary(domain: str, entries: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """构建扩展字典，不修改全局语义库，也不修改调用方传入的数据。"""
    canonical_domain = _canonical_domain(domain)
    if not canonical_domain:
        return {}
    result = copy.deepcopy(ALL_DICTIONARIES.get(canonical_domain, {}))
    for raw_entry in entries:
        entry = copy.deepcopy(raw_entry)
        name = str(entry.pop("name", entry.pop("canonical", ""))).strip()
        if not name:
            continue
        current = copy.deepcopy(result.get(name, {}))
        current["aliases"] = list(dict.fromkeys(_as_alias_list(current.get("aliases")) + _as_alias_list(entry.get("aliases"))))
        current.update({k: v for k, v in entry.items() if k != "aliases"})
        old_codes = _entry_codes(result.get(name, {}))
        new_codes = _entry_codes(entry)
        if old_codes or new_codes:
            current["codes"] = {**old_codes, **new_codes}
        result[name] = current
    return result


# ---------------------------------------------------------------------------
# 文件读取与自定义字典识别
# ---------------------------------------------------------------------------


def _read_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def _read_csv_rows(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    text = _read_text(path)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel_tab if path.suffix.lower() == ".txt" and "\t" in sample else csv.excel
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    fields = [str(f).strip() for f in (reader.fieldnames or []) if f is not None]
    return [dict(row) for row in reader], fields


def _read_xlsx_rows(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("读取 XLSX 需要安装 openpyxl：pip install openpyxl") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return [], []
    fields = [str(v).strip() if v is not None else "" for v in header_row]
    rows = []
    for values in iterator:
        rows.append({fields[i]: values[i] if i < len(values) else None for i in range(len(fields)) if fields[i]})
    return rows, [f for f in fields if f]


def _dictionary_columns(fieldnames: Iterable[str]) -> dict[str, str] | None:
    normalized = {_normalize_token(f): f for f in fieldnames}
    aliases = {
        "domain": ("domain", "语义域", "类别", "字典类别"),
        "canonical": ("canonical", "规范名", "标准名称", "规范名称"),
        "aliases": ("aliases", "alias", "别名", "同义词"),
        "project_code": ("projectcode", "项目编码", "内部编码", "标准编码"),
        "symbol": ("symbol", "地质符号", "代号"),
        "description": ("description", "描述", "说明"),
        "category": ("category", "一级分类", "类型"),
        "subcategory": ("subcategory", "二级分类", "子类"),
        "unit": ("unit", "单位"),
    }
    found: dict[str, str] = {}
    for key, candidates in aliases.items():
        for candidate in candidates:
            original = normalized.get(_normalize_token(candidate))
            if original:
                found[key] = original
                break
    if "domain" in found and "canonical" in found:
        return found
    return None


def _rows_to_custom_dictionary(rows: list[dict[str, Any]], columns: dict[str, str]) -> dict[str, dict[str, dict[str, Any]]]:
    result: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    for row in rows:
        domain = _canonical_domain(str(row.get(columns["domain"], "")))
        canonical = str(row.get(columns["canonical"], "")).strip()
        if not domain or not canonical:
            continue
        entry: dict[str, Any] = {
            "aliases": _as_alias_list(row.get(columns.get("aliases", ""), "")),
            "description": str(row.get(columns.get("description", ""), "")).strip(),
        }
        codes: dict[str, str] = {}
        if columns.get("project_code") and row.get(columns["project_code"]):
            codes["PROJECT"] = str(row[columns["project_code"]]).strip()
        if columns.get("symbol") and row.get(columns["symbol"]):
            codes["SYMBOL"] = str(row[columns["symbol"]]).strip()
        if codes:
            entry["codes"] = codes
        for key in ("category", "subcategory", "unit"):
            column = columns.get(key)
            if column and row.get(column) not in (None, ""):
                entry[key] = str(row[column]).strip()
        result[domain][canonical] = entry
    return dict(result)


def _load_dictionary_file(path: Path) -> dict[str, dict[str, dict[str, Any]]] | None:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        rows, fields = _read_csv_rows(path)
        columns = _dictionary_columns(fields)
        return _rows_to_custom_dictionary(rows, columns) if columns else None
    if suffix == ".xlsx":
        rows, fields = _read_xlsx_rows(path)
        columns = _dictionary_columns(fields)
        return _rows_to_custom_dictionary(rows, columns) if columns else None
    if suffix == ".json":
        data = json.loads(_read_text(path))
        if isinstance(data, dict) and isinstance(data.get("domains"), dict):
            return {str(k): v for k, v in data["domains"].items() if isinstance(v, dict)}
    return None


# ---------------------------------------------------------------------------
# 字段识别
# ---------------------------------------------------------------------------

_EXACT_FIELD_DOMAIN = {
    "地层": "地层", "地层名称": "地层", "地层代号": "地层", "时代": "地层", "组名": "地层",
    "岩性": "岩性", "岩性名称": "岩性", "主要岩性": "岩性", "岩石名称": "岩性", "土层名称": "岩性",
    "构造": "构造", "构造类型": "构造", "断层类型": "构造", "结构面类型": "构造",
    "地下水": "地下水", "地下水类型": "地下水", "含水类型": "地下水", "出水状态": "地下水",
    "不良地质": "不良地质", "灾害类型": "不良地质", "地质灾害": "不良地质",
    "围岩等级": "围岩等级", "围岩级别": "围岩等级", "围岩类别": "围岩等级",
    "工程地质指标": "工程地质指标", "指标名称": "工程地质指标", "试验指标": "工程地质指标",
    "含水率": "工程地质指标", "渗透系数": "工程地质指标", "抗压强度": "工程地质指标",
    "单轴抗压强度": "工程地质指标", "弹性模量": "工程地质指标", "泊松比": "工程地质指标",
    "RQD": "工程地质指标", "RMR": "工程地质指标", "BQ": "工程地质指标", "波速": "工程地质指标",
}

_FIELD_SIGNALS: dict[str, tuple[str, ...]] = {
    "工程地质指标": ("强度", "模量", "密度", "重度", "含水率", "孔隙", "rq", "rmr", "bq", "波速", "倾角", "倾向", "摩擦角", "黏聚力", "承载力", "贯入", "锥尖", "指标"),
    "围岩等级": ("围岩等级", "围岩级别", "围岩类别", "rockclass", "classification"),
    "不良地质": ("不良地质", "灾害", "hazard", "adverse", "滑坡", "泥石流", "塌陷", "采空", "岩爆"),
    "地下水": ("地下水类型", "含水类型", "出水状态", "groundwater", "watertype", "水文类型", "涌水状态"),
    "构造": ("构造类型", "断层类型", "节理类型", "structure", "faulttype", "jointtype", "褶皱类型", "接触关系"),
    "岩性": ("岩性", "岩石名称", "主要岩性", "lithology", "rocktype", "soiltype", "土层名称"),
    "地层": ("地层名称", "地层代号", "strata", "formation", "period", "时代", "组名"),
}


def _maybe_semantic_field(field_name: str) -> str | None:
    name = str(field_name or "").strip()
    if not name:
        return None
    if name in _EXACT_FIELD_DOMAIN:
        return _EXACT_FIELD_DOMAIN[name]
    token = _normalize_token(name)
    if token in {"项目组", "系统", "分类", "类别", "名称", "编码", "id", "code"}:
        return None
    if token.endswith(("指标值", "数值", "测值", "单位", "备注", "说明")) or token in {"值", "value"}:
        return None
    for domain, signals in _FIELD_SIGNALS.items():
        if any(_normalize_token(signal) in token for signal in signals):
            return domain
    return None


def _infer_domain_from_values(values: Iterable[Any], dictionaries: dict[str, dict[str, dict[str, Any]]]) -> str | None:
    sample = [str(v).strip() for v in values if _normalize_token(v) not in _NULL_VALUES][:30]
    if not sample:
        return None
    scores: dict[str, int] = {d: 0 for d in DOMAIN_ORDER}
    for value in sample:
        for domain in DOMAIN_ORDER:
            if normalize_name(value, domain, dictionaries=dictionaries)["confidence"] >= 0.9:
                scores[domain] += 1
    best_domain, best_score = max(scores.items(), key=lambda item: item[1])
    return best_domain if best_score >= max(2, len(sample) // 2) else None


# ---------------------------------------------------------------------------
# 报告、分析与导出
# ---------------------------------------------------------------------------


def _new_semantic_report() -> dict[str, Any]:
    return {
        "schema_version": "2.0",
        "run_id": f"SEM-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8].upper()}",
        "checked_at": datetime.now().isoformat(timespec="seconds"),
        "files": [],
        "mappings": [],
        "conflicts": [],
        "issues": [],
        "relations": [],
        "normalized_files": [],
        "total_terms": 0,
        "domain_stats": {},
        "library_stats": {},
        "custom_dictionary_terms": 0,
        "output_artifacts": [],
    }


def _add_semantic_issue(
    report: dict[str, Any],
    file_name: str,
    severity: str,
    message: str,
    field_name: str | None = None,
    suggestion: str = "",
    code: str = "SEMANTIC_MAPPING",
    location: str = "",
) -> None:
    report["issues"].append({
        "issue_id": f"S-{len(report['issues']) + 1:04d}",
        "code": code,
        "file": file_name,
        "field": field_name,
        "location": location or (f"字段 {field_name}" if field_name else "文件级"),
        "severity": severity,
        "message": message,
        "suggestion": suggestion,
    })


def _register_mapping(
    report: dict[str, Any],
    path: Path,
    field_name: str,
    domain: str,
    original: str,
    norm: dict[str, Any],
    conflict_strategy: str,
    target_code_system: str,
) -> None:
    codes = norm.get("codes", {}) if isinstance(norm.get("codes"), dict) else {}
    mapping = {
        "file": path.name,
        "field": field_name,
        "domain": domain,
        "original": original,
        "canonical": norm.get("canonical", original),
        "matched_by": norm.get("matched", "未匹配"),
        "confidence": float(norm.get("confidence", 0) or 0),
        "source": norm.get("source", "original"),
        "project_code": codes.get("PROJECT", ""),
        "symbol": codes.get("SYMBOL", ""),
        # 兼容旧界面字段名；实际编码体系由 code_system 明确标识
        "standard_code": codes.get("PROJECT", ""),
        "target_code": codes.get(target_code_system, ""),
        "code_system": target_code_system,
        "category": norm.get("category", ""),
        "subcategory": norm.get("subcategory", ""),
        "unit": norm.get("unit", ""),
        "candidates": norm.get("candidates", []),
    }
    report["mappings"].append(mapping)

    candidates = norm.get("candidates", [])
    if isinstance(candidates, list) and len(candidates) > 1:
        resolution = {
            "人工复核": "人工复核",
            "规范覆盖": f"采用 {norm.get('canonical', original)}",
            "保留原始": "保留原始",
        }.get(conflict_strategy, "人工复核")
        report["conflicts"].append({
            "original": original,
            "field": field_name,
            "file": path.name,
            "domain": domain,
            "canonicals": candidates,
            "resolution": resolution,
            "message": f"同一术语命中多个候选：{'、'.join(candidates)}",
        })

    for relation in norm.get("relations", []) if isinstance(norm.get("relations"), list) else []:
        report["relations"].append({
            "file": path.name,
            "field": field_name,
            "source_domain": domain,
            "source_term": norm.get("canonical", original),
            **copy.deepcopy(relation),
        })


def _semantic_fields_for_rows(
    rows: list[dict[str, Any]],
    fieldnames: list[str],
    dictionaries: dict[str, dict[str, dict[str, Any]]],
    field_domain_map: dict[str, str] | None,
    allowed_domains: set[str],
) -> dict[str, str]:
    result: dict[str, str] = {}
    manual = field_domain_map or {}
    for field in fieldnames:
        domain = _canonical_domain(manual.get(field)) if field in manual else _maybe_semantic_field(field)
        if not domain:
            values = [row.get(field) for row in rows[:100]]
            domain = _infer_domain_from_values(values, dictionaries)
        if domain and domain in allowed_domains:
            result[field] = domain
    return result


def _normalize_rows(
    path: Path,
    rows: list[dict[str, Any]],
    fieldnames: list[str],
    report: dict[str, Any],
    dictionaries: dict[str, dict[str, dict[str, Any]]],
    domains: set[str],
    synonym_strategy: str,
    conflict_strategy: str,
    target_code_system: str,
    field_domain_map: dict[str, str] | None,
) -> tuple[list[dict[str, Any]], list[str], dict[str, str]]:
    semantic_fields = _semantic_fields_for_rows(rows, fieldnames, dictionaries, field_domain_map, domains)
    if not semantic_fields:
        _add_semantic_issue(
            report, path.name, "提示",
            "未识别到地层、岩性、构造、地下水、不良地质、围岩等级或工程地质指标字段。",
            suggestion="可规范字段名，或通过 field_domain_map 手动指定字段与语义域。",
        )
        return copy.deepcopy(rows), list(fieldnames), semantic_fields

    normalized_rows = copy.deepcopy(rows)
    output_fields = list(fieldnames)
    unique_by_field: dict[str, set[str]] = {f: set() for f in semantic_fields}
    for row in rows:
        for field in semantic_fields:
            raw = str(row.get(field, "") or "").strip()
            if _normalize_token(raw) not in _NULL_VALUES:
                unique_by_field[field].add(raw)

    cache: dict[tuple[str, str], dict[str, Any]] = {}
    for field, domain in semantic_fields.items():
        values = sorted(unique_by_field[field])[:500]
        report["total_terms"] += len(values)
        for value in values:
            norm = normalize_name(value, domain, strategy=synonym_strategy, dictionaries=dictionaries)
            if len(norm.get("candidates", [])) > 1 and conflict_strategy in {"人工复核", "保留原始"}:
                norm = copy.deepcopy(norm)
                norm["canonical"] = value
                norm["codes"] = {}
                norm["matched"] = "冲突待复核" if conflict_strategy == "人工复核" else "冲突保留原始"
            cache[(field, value)] = norm
            _register_mapping(report, path, field, domain, value, norm, conflict_strategy, target_code_system)
            if float(norm.get("confidence", 0) or 0) == 0:
                _add_semantic_issue(
                    report, path.name, "警告",
                    f"“{value}”（字段 {field}）在{domain}语义库中未匹配。",
                    field_name=field,
                    suggestion="请在项目语义字典中补充规范名、别名或项目编码。",
                )

        for suffix in ("规范名", "项目编码", "地质符号", "目标编码", "编码体系", "匹配方式", "置信度"):
            new_field = f"{field}_{suffix}"
            if new_field not in output_fields:
                output_fields.append(new_field)

    for row in normalized_rows:
        for field, domain in semantic_fields.items():
            raw = str(row.get(field, "") or "").strip()
            if _normalize_token(raw) in _NULL_VALUES:
                continue
            norm = cache.get((field, raw)) or normalize_name(raw, domain, strategy=synonym_strategy, dictionaries=dictionaries)
            codes = norm.get("codes", {}) if isinstance(norm.get("codes"), dict) else {}
            row[f"{field}_规范名"] = norm.get("canonical", raw)
            row[f"{field}_项目编码"] = codes.get("PROJECT", "")
            row[f"{field}_地质符号"] = codes.get("SYMBOL", "")
            row[f"{field}_目标编码"] = codes.get(target_code_system, "")
            row[f"{field}_编码体系"] = target_code_system
            row[f"{field}_匹配方式"] = norm.get("matched", "未匹配")
            row[f"{field}_置信度"] = norm.get("confidence", 0)
    return normalized_rows, output_fields, semantic_fields


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    with temporary.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    temporary.replace(path)


def _write_xlsx(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ValueError("导出 XLSX 需要安装 openpyxl：pip install openpyxl") from exc
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "语义归一结果"
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(field, "") for field in fieldnames])
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    wb.save(temporary)
    temporary.replace(path)


def _analyze_tabular_file(
    path: Path,
    report: dict[str, Any],
    entry: dict[str, Any],
    dictionaries: dict[str, dict[str, dict[str, Any]]],
    domains: set[str],
    synonym_strategy: str,
    conflict_strategy: str,
    target_code_system: str,
    field_domain_map: dict[str, str] | None,
    output_dir: Path | None,
) -> None:
    if path.suffix.lower() == ".xlsx":
        rows, fields = _read_xlsx_rows(path)
    else:
        rows, fields = _read_csv_rows(path)
    entry["row_count"] = len(rows)
    entry["field_count"] = len(fields)
    if not fields:
        _add_semantic_issue(report, path.name, "严重", "表格缺少有效表头。")
        return
    normalized_rows, output_fields, semantic_fields = _normalize_rows(
        path, rows, fields, report, dictionaries, domains,
        synonym_strategy, conflict_strategy, target_code_system, field_domain_map,
    )
    entry["semantic_fields"] = semantic_fields
    if output_dir and normalized_rows and semantic_fields:
        suffix = ".xlsx" if path.suffix.lower() == ".xlsx" else ".csv"
        output = output_dir / f"{path.stem}_semantic_normalized{suffix}"
        if suffix == ".xlsx":
            _write_xlsx(output, normalized_rows, output_fields)
        else:
            _write_csv(output, normalized_rows, output_fields)
        report["normalized_files"].append(str(output.resolve()))
        entry["normalized_output"] = str(output.resolve())
        _register_output_artifact(report, path, output)


def _json_feature_properties(data: Any) -> tuple[list[dict[str, Any]], str]:
    if isinstance(data, dict) and data.get("type") == "FeatureCollection":
        return [f.get("properties", {}) for f in data.get("features", []) if isinstance(f, dict) and isinstance(f.get("properties"), dict)], "FeatureCollection"
    if isinstance(data, dict) and data.get("type") == "Feature" and isinstance(data.get("properties"), dict):
        return [data["properties"]], "Feature"
    if isinstance(data, list) and all(isinstance(item, dict) for item in data):
        return data, "List"
    if isinstance(data, dict):
        return [data], "Object"
    return [], "Unknown"


def _analyze_json_file(
    path: Path,
    report: dict[str, Any],
    entry: dict[str, Any],
    dictionaries: dict[str, dict[str, dict[str, Any]]],
    domains: set[str],
    synonym_strategy: str,
    conflict_strategy: str,
    target_code_system: str,
    field_domain_map: dict[str, str] | None,
    output_dir: Path | None,
) -> None:
    data = json.loads(_read_text(path))
    rows, json_type = _json_feature_properties(data)
    entry["json_type"] = json_type
    entry["feature_count"] = len(rows)
    if not rows:
        _add_semantic_issue(report, path.name, "提示", "JSON/GeoJSON 中未找到可解析的属性记录。")
        return
    fields = list(dict.fromkeys(str(key) for row in rows for key in row.keys()))
    normalized_rows, _, semantic_fields = _normalize_rows(
        path, rows, fields, report, dictionaries, domains,
        synonym_strategy, conflict_strategy, target_code_system, field_domain_map,
    )
    entry["semantic_fields"] = semantic_fields
    if not output_dir or not semantic_fields:
        return

    if json_type == "FeatureCollection":
        normalized_data = copy.deepcopy(data)
        idx = 0
        for feature in normalized_data.get("features", []):
            if isinstance(feature, dict) and isinstance(feature.get("properties"), dict):
                feature["properties"] = normalized_rows[idx]
                idx += 1
    elif json_type == "Feature":
        normalized_data = copy.deepcopy(data)
        normalized_data["properties"] = normalized_rows[0]
    elif json_type == "List":
        normalized_data = normalized_rows
    else:
        normalized_data = normalized_rows[0]
    output = output_dir / f"{path.stem}_semantic_normalized{path.suffix.lower()}"
    output.parent.mkdir(parents=True, exist_ok=True)
    _atomic_text(output, json.dumps(normalized_data, ensure_ascii=False, indent=2))
    report["normalized_files"].append(str(output.resolve()))
    entry["normalized_output"] = str(output.resolve())
    _register_output_artifact(report, path, output)


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _atomic_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    temporary.write_text(content, encoding="utf-8")
    temporary.replace(path)


def _register_output_artifact(report: dict[str, Any], source: Path, output: Path) -> None:
    report["output_artifacts"].append({
        "source": str(source.resolve()),
        "source_sha256": _file_sha256(source),
        "output": str(output.resolve()),
        "output_sha256": _file_sha256(output),
        "size_bytes": output.stat().st_size,
    })


def audit_semantic_library(
    dictionaries: dict[str, dict[str, dict[str, Any]]] | None = None,
) -> dict[str, Any]:
    """检查语义库内重复规范词、别名冲突、编码冲突和缺失项目编码。"""
    dictionaries = dictionaries or ALL_DICTIONARIES
    alias_index: dict[tuple[str, str], set[str]] = defaultdict(set)
    code_index: dict[tuple[str, str, str], set[str]] = defaultdict(set)
    missing_project_codes: list[dict[str, str]] = []
    for domain, entries in dictionaries.items():
        for canonical, entry in entries.items():
            for token in [canonical, *_as_alias_list(entry.get("aliases"))]:
                normalized = _normalize_token(token)
                if normalized:
                    alias_index[(domain, normalized)].add(canonical)
            codes = _entry_codes(entry)
            if not codes.get("PROJECT"):
                missing_project_codes.append({"domain": domain, "canonical": canonical})
            for system, code in codes.items():
                code_index[(domain, system, _normalize_token(code))].add(canonical)
    alias_conflicts = [
        {"domain": domain, "token": token, "canonicals": sorted(canonicals)}
        for (domain, token), canonicals in alias_index.items() if len(canonicals) > 1
    ]
    code_conflicts = [
        {"domain": domain, "system": system, "code": code, "canonicals": sorted(canonicals)}
        for (domain, system, code), canonicals in code_index.items() if code and len(canonicals) > 1
    ]
    return {
        "valid": not alias_conflicts and not code_conflicts,
        "alias_conflicts": alias_conflicts,
        "code_conflicts": code_conflicts,
        "missing_project_codes": missing_project_codes,
        "stats": semantic_library_stats(dictionaries),
    }


def _write_semantic_manifest(output_dir: Path, report: dict[str, Any]) -> str:
    manifest = output_dir / "semantic_conversion_manifest.json"
    payload = copy.deepcopy(report)
    payload["manifest_path"] = str(manifest.resolve())
    _atomic_text(manifest, json.dumps(payload, ensure_ascii=False, indent=2))
    return str(manifest.resolve())


def _semantic_grade(score: int) -> str:
    if score >= 95:
        return "优秀"
    if score >= 85:
        return "良好"
    if score >= 70:
        return "合格"
    return "需复核"


def _finalize_semantic(report: dict[str, Any]) -> dict[str, Any]:
    total = int(report.get("total_terms", 0) or 0)
    mapped = sum(1 for m in report.get("mappings", []) if float(m.get("confidence", 0) or 0) > 0)
    unmapped = [m for m in report.get("mappings", []) if float(m.get("confidence", 0) or 0) == 0]
    report["coverage_ratio"] = round(mapped / total, 4) if total else 0.0
    report["unmapped_terms"] = unmapped

    for domain in DOMAIN_ORDER:
        items = [m for m in report.get("mappings", []) if m.get("domain") == domain]
        domain_mapped = sum(1 for m in items if float(m.get("confidence", 0) or 0) > 0)
        report["domain_stats"][domain] = {
            "total": len(items),
            "mapped": domain_mapped,
            "unmapped": len(items) - domain_mapped,
            "coverage": round(domain_mapped / len(items), 4) if items else 0.0,
        }

    severity_cost = {"严重": 12, "警告": 3, "提示": 1}
    issue_penalty = sum(severity_cost.get(i.get("severity", ""), 1) for i in report.get("issues", []))
    coverage_penalty = round((1 - report["coverage_ratio"]) * 35) if total else 10
    conflict_penalty = min(15, len(report.get("conflicts", [])) * 3)
    score = 100 - issue_penalty - coverage_penalty - conflict_penalty
    report["score"] = max(0, min(100, int(score)))
    report["grade"] = _semantic_grade(report["score"])
    report["issue_count"] = len(report.get("issues", []))
    report["file_count"] = len(report.get("files", []))
    return report


def build_semantic_report(
    file_paths: list[str],
    target_domains: list[str] | None = None,
    synonym_strategy: str = "规范优先",
    conflict_resolution: str = "人工复核",
    custom_dictionaries: dict[str, dict[str, dict[str, Any]]] | None = None,
    field_domain_map: dict[str, str] | None = None,
    output_dir: str | Path | None = None,
    library_path: str | Path | None = None,
    target_code_system: str = "PROJECT",
) -> dict[str, Any]:
    """执行字典加载、名称归一、编码映射、语义关联和标准化成果导出。"""
    report = _new_semantic_report()
    target_code_system = str(target_code_system or "PROJECT").strip().upper()
    if target_code_system not in ENCODING_SYSTEMS:
        target_code_system = "PROJECT"
    requested = target_domains or list(DOMAIN_ORDER)
    domains = {_canonical_domain(d) for d in requested}
    domains = {d for d in domains if d}
    if not domains:
        domains = set(DOMAIN_ORDER)

    base_library = load_semantic_library(library_path) if library_path else {"metadata": LIBRARY_METADATA, "domains": ALL_DICTIONARIES}
    working = _merge_dictionary_sets(
        {domain: copy.deepcopy(base_library.get("domains", {}).get(domain, {})) for domain in DOMAIN_ORDER},
        custom_dictionaries,
    )
    report["parameters"] = {
        "target_domains": [d for d in DOMAIN_ORDER if d in domains],
        "synonym_strategy": synonym_strategy,
        "conflict_resolution": conflict_resolution,
        "target_code_system": target_code_system,
        "output_dir": str(output_dir) if output_dir else "",
    }
    report["library_metadata"] = copy.deepcopy(base_library.get("metadata", {}))

    if not file_paths:
        _add_semantic_issue(report, "未选择文件", "警告", "请先导入地质属性数据、项目编码表或语义字典。")
        report["library_stats"] = semantic_library_stats(working, base_library.get("metadata", {}))
        report["library_audit"] = audit_semantic_library(working)
        return _finalize_semantic(report)

    paths = [Path(p) for p in file_paths]
    entries_by_path: dict[Path, dict[str, Any]] = {}
    # 第一遍：登记文件并优先加载项目自定义字典
    for path in paths:
        entry = {"path": str(path), "name": path.name, "type": path.suffix.lower() or "unknown", "status": "待分析"}
        report["files"].append(entry)
        entries_by_path[path] = entry
        if not path.exists():
            entry["status"] = "文件缺失"
            _add_semantic_issue(report, path.name, "严重", "文件不存在或路径不可访问。")
            continue
        entry["source_sha256"] = _file_sha256(path)
        try:
            custom_from_file = _load_dictionary_file(path)
        except (OSError, ValueError, json.JSONDecodeError, csv.Error) as exc:
            custom_from_file = None
            entry["dictionary_load_error"] = str(exc)
        if custom_from_file:
            working = _merge_dictionary_sets(working, custom_from_file)
            added = sum(len(v) for v in custom_from_file.values())
            report["custom_dictionary_terms"] += added
            entry["status"] = "已加载项目字典"
            entry["dictionary_terms"] = added
            entry["is_dictionary"] = True

    report["library_stats"] = semantic_library_stats(working, base_library.get("metadata", {}))
    report["library_audit"] = audit_semantic_library(working)
    out_dir = Path(output_dir) if output_dir else None
    if out_dir:
        out_dir.mkdir(parents=True, exist_ok=True)

    # 第二遍：分析业务数据并输出规范化成果
    for path in paths:
        entry = entries_by_path[path]
        if entry.get("status") == "文件缺失" or entry.get("is_dictionary"):
            continue
        try:
            suffix = path.suffix.lower()
            if suffix in {".csv", ".txt", ".xlsx"}:
                _analyze_tabular_file(
                    path, report, entry, working, domains,
                    synonym_strategy, conflict_resolution, target_code_system, field_domain_map, out_dir,
                )
                entry["status"] = "已转换"
            elif suffix in {".json", ".geojson"}:
                _analyze_json_file(
                    path, report, entry, working, domains,
                    synonym_strategy, conflict_resolution, target_code_system, field_domain_map, out_dir,
                )
                entry["status"] = "已转换"
            else:
                entry["status"] = "格式未识别"
                _add_semantic_issue(report, path.name, "提示", f"{suffix or '未知'} 格式暂未内置语义解析器。")
        except (OSError, UnicodeDecodeError, json.JSONDecodeError, csv.Error, ValueError) as exc:
            entry["status"] = "解析失败"
            _add_semantic_issue(report, path.name, "严重", f"文件解析失败：{exc}")

    finalized = _finalize_semantic(report)
    if out_dir:
        finalized["manifest_path"] = _write_semantic_manifest(out_dir, finalized)
    return finalized


def export_normalized_table(
    file_path: str,
    output_path: str,
    field_domain_map: dict[str, str] | None = None,
    target_domains: list[str] | None = None,
    synonym_strategy: str = "规范优先",
    custom_dictionaries: dict[str, dict[str, dict[str, Any]]] | None = None,
) -> dict[str, Any]:
    """兼容旧接口：规范化单个表格并导出。"""
    source = Path(file_path)
    output = Path(output_path)
    report = build_semantic_report(
        [str(source)],
        target_domains=target_domains,
        synonym_strategy=synonym_strategy,
        custom_dictionaries=custom_dictionaries,
        field_domain_map=field_domain_map,
        output_dir=output.parent,
    )
    generated = report.get("normalized_files", [])
    if generated:
        generated_path = Path(generated[0])
        if generated_path.resolve() != output.resolve():
            output.parent.mkdir(parents=True, exist_ok=True)
            output.write_bytes(generated_path.read_bytes())
    return {
        "output": str(output.resolve()) if output.exists() else None,
        "rows": next((f.get("row_count", 0) for f in report.get("files", []) if f.get("name") == source.name), 0),
        "fields_normalized": next((list((f.get("semantic_fields") or {}).keys()) for f in report.get("files", []) if f.get("name") == source.name), []),
        "unmatched": list(dict.fromkeys(m.get("original", "") for m in report.get("unmapped_terms", []))),
        "report": report,
    }


__all__ = [
    "ALL_DICTIONARIES",
    "CATEGORY_LABELS",
    "DOMAIN_ORDER",
    "ENCODING_SYSTEMS",
    "LIBRARY_METADATA",
    "STRATA_DICTIONARY",
    "LITHOLOGY_DICTIONARY",
    "STRUCTURE_DICTIONARY",
    "GROUNDWATER_DICTIONARY",
    "ADVERSE_GEO_DICTIONARY",
    "ROCK_CLASS_DICTIONARY",
    "ENGINEERING_INDICATORS_DICTIONARY",
    "build_semantic_dictionary",
    "build_semantic_report",
    "audit_semantic_library",
    "export_normalized_table",
    "load_semantic_library",
    "map_encoding",
    "normalize_name",
    "semantic_library_stats",
]
