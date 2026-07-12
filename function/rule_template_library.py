from __future__ import annotations

import ast
import copy
import csv
import hashlib
import json
import math
import operator
import re
import uuid
from dataclasses import asdict, dataclass, field, fields
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable


# ============================================================================
# 基础常量
# ============================================================================

SCHEMA_VERSION = "2.0"
BUILTIN_CREATED_AT = "2026-07-11T00:00:00"

PROJECT_TYPES = (
    "隧道工程",
    "边坡工程",
    "地基与基础工程",
    "矿山工程",
    "水利水电工程",
    "城市地质调查",
    "区域地质调查",
    "环境地质评价",
    "通用工程",
)

DATA_SOURCES = (
    "钻孔数据",
    "地质图",
    "物探数据",
    "遥感影像",
    "点云数据",
    "地质剖面",
    "属性表",
    "网格模型",
    "通用数据",
)

TARGET_PLATFORMS = ("GIS平台", "三维建模软件", "BIM平台", "数值模拟软件")

COMMON_INPUT_FORMATS = (
    "CSV", "XLSX", "JSON", "GeoJSON", "SHP", "GPKG", "DXF", "DWG",
    "OBJ", "STL", "PLY", "LAS", "LAZ", "TIFF", "GeoTIFF", "VTK",
    "VTU", "INP", "IFC", "TXT",
)

MISSING_MARKERS = {"", "-", "--", "none", "null", "nan", "n/a", "无", "缺失"}


# ============================================================================
# 数据结构
# ============================================================================

@dataclass
class FieldMappingRule:
    """字段映射规则：源字段（及别名）→ 标准目标字段。"""

    source_field: str
    target_field: str
    aliases: list[str] = field(default_factory=list)
    data_type: str = "string"              # string / integer / float / boolean / date
    transform: str = ""                    # 受限安全表达式，例如 float(value) / value * 0.001
    required: bool = False
    default_value: Any = None
    source_unit: str = ""
    target_unit: str = ""
    description: str = ""


@dataclass
class CoordinateRule:
    """坐标与高程基准规则。真实投影转换需由 pyproj/GDAL 后端执行。"""

    source_crs: str = ""
    target_crs: str = ""
    source_epsg: str = ""
    target_epsg: str = ""
    source_datum: str = ""
    target_datum: str = ""
    axis_order: str = "XY"
    horizontal_unit: str = "m"
    vertical_unit: str = "m"
    offset_x: float = 0.0
    offset_y: float = 0.0
    offset_z: float = 0.0
    scale_factor: float = 1.0
    precision: int = 6
    description: str = ""


@dataclass
class AttributeMappingRule:
    """属性枚举值映射规则。"""

    source_field: str
    target_field: str
    value_map: dict[str, str] = field(default_factory=dict)
    default_value: str = ""
    case_sensitive: bool = False
    strip_value: bool = True
    description: str = ""


@dataclass
class OutputRule:
    """成果输出规则。"""

    target_format: str = "CSV"
    target_platform: str = "GIS平台"
    encoding: str = "UTF-8"
    include_normals: bool = True
    include_textures: bool = False
    include_properties: bool = True
    coordinate_precision: int = 6
    layer_name: str = ""
    file_name_pattern: str = "{template}_{date}"
    platform_options: dict[str, Any] = field(default_factory=dict)
    description: str = ""


@dataclass
class QualityCheckRule:
    """质量检查规则。"""

    check_geometry_closure: bool = True
    check_topology: bool = True
    check_attribute_completeness: bool = True
    check_consistency: bool = True
    check_coordinate_precision: bool = True
    check_duplicates: bool = True
    check_units: bool = True
    closure_tolerance: float = 0.001
    outlier_sigma: float = 3.0
    min_face_quality: float = 0.3
    max_missing_ratio: float = 0.1
    required_fields: list[str] = field(default_factory=list)
    unique_fields: list[str] = field(default_factory=list)
    numeric_ranges: dict[str, list[float | None]] = field(default_factory=dict)
    allowed_values: dict[str, list[str]] = field(default_factory=dict)
    description: str = ""


@dataclass
class RuleTemplate:
    """完整转换规则模板。"""

    name: str
    version: str = "1.0.0"
    schema_version: str = SCHEMA_VERSION
    project_type: str = "通用工程"
    data_source: str = "通用数据"
    input_formats: list[str] = field(default_factory=list)
    description: str = ""
    author: str = ""
    created_at: str = ""
    updated_at: str = ""
    tags: list[str] = field(default_factory=list)
    enabled: bool = True

    field_mappings: list[FieldMappingRule] = field(default_factory=list)
    coordinate_rule: CoordinateRule | None = None
    attribute_mappings: list[AttributeMappingRule] = field(default_factory=list)
    output_rule: OutputRule | None = None
    quality_check_rule: QualityCheckRule | None = None

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "RuleTemplate":
        clean = dict(data)
        clean.setdefault("schema_version", SCHEMA_VERSION)
        clean.setdefault("input_formats", [])
        clean.setdefault("enabled", True)
        clean["field_mappings"] = [
            _dataclass_from_dict(FieldMappingRule, item)
            for item in clean.get("field_mappings", [])
        ]
        clean["coordinate_rule"] = (
            _dataclass_from_dict(CoordinateRule, clean["coordinate_rule"])
            if clean.get("coordinate_rule") else None
        )
        clean["attribute_mappings"] = [
            _dataclass_from_dict(AttributeMappingRule, item)
            for item in clean.get("attribute_mappings", [])
        ]
        clean["output_rule"] = (
            _dataclass_from_dict(OutputRule, clean["output_rule"])
            if clean.get("output_rule") else None
        )
        clean["quality_check_rule"] = (
            _dataclass_from_dict(QualityCheckRule, clean["quality_check_rule"])
            if clean.get("quality_check_rule") else None
        )
        allowed = {item.name for item in fields(cls)}
        return cls(**{key: value for key, value in clean.items() if key in allowed})


@dataclass
class TemplateValidationResult:
    valid: bool
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# ============================================================================
# 工具函数
# ============================================================================

def _dataclass_from_dict(cls: type[Any], data: dict[str, Any]) -> Any:
    allowed = {item.name for item in fields(cls)}
    return cls(**{key: value for key, value in data.items() if key in allowed})


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _atomic_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def template_fingerprint(template: RuleTemplate) -> str:
    payload = template.to_dict()
    payload.pop("created_at", None)
    payload.pop("updated_at", None)
    canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _fm(
    source: str,
    target: str,
    *,
    aliases: Iterable[str] = (),
    data_type: str = "string",
    transform: str = "",
    required: bool = False,
    default: Any = None,
    source_unit: str = "",
    target_unit: str = "",
    description: str = "",
) -> FieldMappingRule:
    return FieldMappingRule(
        source_field=source,
        target_field=target,
        aliases=list(aliases),
        data_type=data_type,
        transform=transform,
        required=required,
        default_value=default,
        source_unit=source_unit,
        target_unit=target_unit,
        description=description,
    )


def _coord(
    source_crs: str = "CGCS2000",
    target_crs: str = "CGCS2000",
    *,
    source_epsg: str = "",
    target_epsg: str = "",
    source_datum: str = "1985国家高程基准",
    target_datum: str = "1985国家高程基准",
    description: str = "",
) -> CoordinateRule:
    return CoordinateRule(
        source_crs=source_crs,
        target_crs=target_crs,
        source_epsg=source_epsg,
        target_epsg=target_epsg,
        source_datum=source_datum,
        target_datum=target_datum,
        description=description,
    )


def _out(
    fmt: str,
    platform: str,
    *,
    precision: int = 6,
    include_normals: bool = True,
    include_textures: bool = False,
    include_properties: bool = True,
    layer_name: str = "",
    description: str = "",
    options: dict[str, Any] | None = None,
) -> OutputRule:
    return OutputRule(
        target_format=fmt,
        target_platform=platform,
        coordinate_precision=precision,
        include_normals=include_normals,
        include_textures=include_textures,
        include_properties=include_properties,
        layer_name=layer_name,
        description=description,
        platform_options=options or {},
    )


def _qc(
    required: Iterable[str] = (),
    *,
    unique: Iterable[str] = (),
    ranges: dict[str, list[float | None]] | None = None,
    allowed: dict[str, list[str]] | None = None,
    max_missing: float = 0.1,
    topology: bool = True,
    closure: bool = True,
    description: str = "",
) -> QualityCheckRule:
    return QualityCheckRule(
        required_fields=list(required),
        unique_fields=list(unique),
        numeric_ranges=ranges or {},
        allowed_values=allowed or {},
        max_missing_ratio=max_missing,
        check_topology=topology,
        check_geometry_closure=closure,
        description=description,
    )


def _template(
    name: str,
    project_type: str,
    data_source: str,
    input_formats: Iterable[str],
    description: str,
    *,
    version: str = "2.0.0",
    tags: Iterable[str] = (),
    fields_: Iterable[FieldMappingRule] = (),
    coordinate: CoordinateRule | None = None,
    attributes: Iterable[AttributeMappingRule] = (),
    output: OutputRule | None = None,
    quality: QualityCheckRule | None = None,
) -> RuleTemplate:
    return RuleTemplate(
        name=name,
        version=version,
        project_type=project_type,
        data_source=data_source,
        input_formats=list(input_formats),
        description=description,
        author="系统预置",
        created_at=BUILTIN_CREATED_AT,
        updated_at=BUILTIN_CREATED_AT,
        tags=list(tags),
        field_mappings=list(fields_),
        coordinate_rule=coordinate,
        attribute_mappings=list(attributes),
        output_rule=output,
        quality_check_rule=quality,
    )


# ============================================================================
# 完整内置规则模板库
# ============================================================================

def _build_builtin_templates() -> list[RuleTemplate]:
    templates: list[RuleTemplate] = []

    templates.append(_template(
        "隧道工程钻孔地质建模模板", "隧道工程", "钻孔数据", ("CSV", "XLSX", "JSON"),
        "用于铁路、公路隧道钻孔、岩性分层、围岩等级和地下水信息标准化。",
        tags=("隧道", "钻孔", "围岩", "地下水"),
        fields_=(
            _fm("钻孔编号", "borehole_id", aliases=("孔号", "钻孔ID", "hole_id"), required=True),
            _fm("X坐标", "x", aliases=("X", "东坐标", "E"), data_type="float", transform="float(value)", required=True, target_unit="m"),
            _fm("Y坐标", "y", aliases=("Y", "北坐标", "N"), data_type="float", transform="float(value)", required=True, target_unit="m"),
            _fm("孔口高程", "z_surface", aliases=("高程", "Z", "孔口标高"), data_type="float", transform="float(value)", required=True, target_unit="m"),
            _fm("孔深", "total_depth", aliases=("钻孔深度", "终孔深度"), data_type="float", transform="float(value)", required=True, target_unit="m"),
            _fm("层顶深度", "layer_top", aliases=("顶深",), data_type="float", transform="float(value)", default=0.0, target_unit="m"),
            _fm("层底深度", "layer_bottom", aliases=("底深", "分层深度"), data_type="float", transform="float(value)", target_unit="m"),
            _fm("岩性描述", "lithology", aliases=("岩性", "岩性名称"), required=True),
            _fm("围岩等级", "rock_class", aliases=("围岩级别",), required=False),
            _fm("RQD值", "rqd", aliases=("RQD",), data_type="float", transform="float(value)"),
            _fm("地下水", "groundwater", aliases=("地下水情况", "水文情况")),
            _fm("风化程度", "weathering_grade", aliases=("风化等级",)),
        ),
        coordinate=_coord(description="默认采用 CGCS2000 平面坐标及 1985 国家高程基准。"),
        attributes=(
            AttributeMappingRule("rock_class", "rock_class_std", {
                "I": "I级围岩", "Ⅰ": "I级围岩", "II": "II级围岩", "Ⅱ": "II级围岩",
                "III": "III级围岩", "Ⅲ": "III级围岩", "IV": "IV级围岩", "Ⅳ": "IV级围岩",
                "V": "V级围岩", "Ⅴ": "V级围岩", "VI": "VI级围岩", "Ⅵ": "VI级围岩",
            }, description="统一罗马数字和中文围岩等级。"),
            AttributeMappingRule("weathering_grade", "weathering_std", {
                "全风化": "全风化", "强风化": "强风化", "中风化": "中等风化",
                "中等风化": "中等风化", "弱风化": "微风化", "微风化": "微风化", "未风化": "未风化",
            }),
        ),
        output=_out("OBJ", "三维建模软件", precision=4, include_properties=True, layer_name="tunnel_boreholes"),
        quality=_qc(
            ("borehole_id", "x", "y", "z_surface", "total_depth", "lithology"),
            unique=("borehole_id+layer_top",),
            ranges={"total_depth": [0, 5000], "rqd": [0, 100], "layer_top": [0, None], "layer_bottom": [0, None]},
            allowed={"rock_class_std": [f"{x}级围岩" for x in ("I", "II", "III", "IV", "V", "VI")]},
            max_missing=0.08,
        ),
    ))

    templates.append(_template(
        "隧道超前地质预报数据模板", "隧道工程", "物探数据", ("CSV", "XLSX", "JSON", "TXT"),
        "用于 TSP、地质雷达、瞬变电磁、超前水平钻探等预报成果的统一编码。",
        tags=("隧道", "TSP", "GPR", "TEM", "超前预报"),
        fields_=(
            _fm("项目编号", "project_id", aliases=("工程编号",), required=True),
            _fm("方法", "method", aliases=("预报方法", "探测方法"), required=True),
            _fm("起始里程", "start_chainage", aliases=("起点里程", "开始里程"), required=True),
            _fm("终止里程", "end_chainage", aliases=("终点里程", "结束里程"), required=True),
            _fm("长度", "length", aliases=("预报长度",), data_type="float", transform="float(value)", target_unit="m"),
            _fm("综合结论", "conclusion", aliases=("结论", "解释结论"), required=True),
            _fm("岩石坚硬程度", "rock_hardness"),
            _fm("岩体完整程度", "rock_integrity"),
            _fm("地下水状态", "groundwater_state"),
            _fm("风险等级", "risk_level", aliases=("风险级别",)),
        ),
        coordinate=CoordinateRule(source_crs="工程里程坐标", target_crs="工程里程坐标", description="保留线路里程体系，并记录与平面坐标的关联。"),
        attributes=(
            AttributeMappingRule("method", "method_std", {
                "tsp": "TSP", "地震波": "TSP", "gpr": "GPR", "地质雷达": "GPR",
                "tem": "TEM", "瞬变电磁": "TEM", "钻探": "超前水平钻探", "超前钻探": "超前水平钻探",
            }),
            AttributeMappingRule("risk_level", "risk_level_std", {
                "低": "低风险", "一般": "一般风险", "中": "中风险", "高": "高风险", "极高": "极高风险",
            }),
        ),
        output=_out("JSON", "GIS平台", precision=3, include_normals=False, layer_name="tunnel_forecast"),
        quality=_qc(("project_id", "method", "start_chainage", "end_chainage", "conclusion"), max_missing=0.05, topology=False, closure=False),
    ))

    templates.append(_template(
        "边坡工程稳定性评价模板", "边坡工程", "地质剖面", ("CSV", "XLSX", "DXF", "GeoJSON"),
        "用于岩质或土质边坡剖面、节理、岩土参数和稳定性等级的标准化。",
        tags=("边坡", "稳定性", "节理", "剖面"),
        fields_=(
            _fm("剖面编号", "section_id", aliases=("断面编号",), required=True),
            _fm("X坐标", "x", aliases=("X",), data_type="float", transform="float(value)", target_unit="m"),
            _fm("Y坐标", "y", aliases=("Y",), data_type="float", transform="float(value)", target_unit="m"),
            _fm("坡高", "slope_height", aliases=("边坡高度",), data_type="float", transform="float(value)", required=True, target_unit="m"),
            _fm("坡度", "slope_angle", aliases=("坡角",), data_type="float", transform="float(value)", required=True, target_unit="°"),
            _fm("坡向", "slope_aspect", data_type="float", transform="float(value)", target_unit="°"),
            _fm("岩性", "lithology", aliases=("岩性名称",), required=True),
            _fm("节理组数", "joint_sets", data_type="integer", transform="int(value)"),
            _fm("黏聚力", "cohesion", aliases=("粘聚力", "c"), data_type="float", transform="float(value)", target_unit="kPa"),
            _fm("内摩擦角", "friction_angle", aliases=("φ",), data_type="float", transform="float(value)", target_unit="°"),
            _fm("地下水位", "water_table", data_type="float", transform="float(value)", target_unit="m"),
            _fm("坡型", "slope_type"),
            _fm("稳定性", "stability"),
        ),
        coordinate=_coord(description="支持 CGCS2000 与工程独立坐标之间的平移配置。"),
        attributes=(
            AttributeMappingRule("slope_type", "slope_type_std", {"顺层": "顺向坡", "顺向坡": "顺向坡", "反倾": "反向坡", "反向坡": "反向坡", "斜向坡": "斜向坡"}),
            AttributeMappingRule("stability", "stability_std", {"stable": "稳定", "稳定": "稳定", "基本稳定": "基本稳定", "欠稳定": "欠稳定", "unstable": "不稳定", "不稳定": "不稳定"}),
        ),
        output=_out("GeoJSON", "GIS平台", layer_name="slope_sections"),
        quality=_qc(("section_id", "slope_height", "slope_angle", "lithology"), unique=("section_id",), ranges={"slope_height": [0, 2000], "slope_angle": [0, 90], "friction_angle": [0, 90]}, max_missing=0.1, closure=False),
    ))

    templates.append(_template(
        "地基与基础工程勘察模板", "地基与基础工程", "钻孔数据", ("CSV", "XLSX", "JSON"),
        "用于房建、桥梁及场地地基勘察钻孔、土层、标贯和承载力数据。",
        tags=("地基", "基础", "勘察", "标贯"),
        fields_=(
            _fm("钻孔编号", "borehole_id", aliases=("孔号",), required=True),
            _fm("X坐标", "x", aliases=("X",), data_type="float", transform="float(value)", required=True),
            _fm("Y坐标", "y", aliases=("Y",), data_type="float", transform="float(value)", required=True),
            _fm("孔口高程", "z_surface", aliases=("孔口标高",), data_type="float", transform="float(value)", required=True),
            _fm("层顶深度", "layer_top", data_type="float", transform="float(value)", default=0.0),
            _fm("层底深度", "layer_bottom", aliases=("分层深度",), data_type="float", transform="float(value)", required=True),
            _fm("土层名称", "soil_name", aliases=("岩土名称", "地层名称"), required=True),
            _fm("标贯击数", "spt_n", aliases=("SPT_N", "N值"), data_type="float", transform="float(value)"),
            _fm("地基承载力", "bearing_capacity", aliases=("承载力特征值", "faK"), data_type="float", transform="float(value)", target_unit="kPa"),
            _fm("压缩模量", "compression_modulus", aliases=("Es",), data_type="float", transform="float(value)", target_unit="MPa"),
            _fm("地下水位", "water_table", data_type="float", transform="float(value)", target_unit="m"),
        ),
        coordinate=_coord(),
        attributes=(
            AttributeMappingRule("soil_name", "soil_class_std", {"黏土": "黏性土", "粘土": "黏性土", "粉质黏土": "黏性土", "砂土": "砂土", "粉砂": "砂土", "卵石土": "碎石土", "填土": "人工填土"}, default_value="其他土"),
        ),
        output=_out("GeoJSON", "GIS平台", layer_name="foundation_boreholes"),
        quality=_qc(("borehole_id", "x", "y", "z_surface", "layer_bottom", "soil_name"), ranges={"spt_n": [0, 200], "bearing_capacity": [0, 10000]}, max_missing=0.08),
    ))

    templates.append(_template(
        "矿山三维地质建模模板", "矿山工程", "钻孔数据", ("CSV", "XLSX", "JSON", "DXF"),
        "用于矿山勘探钻孔、矿体品位、矿层边界和资源量建模。",
        tags=("矿山", "矿体", "品位", "资源量"),
        fields_=(
            _fm("钻孔编号", "borehole_id", aliases=("孔号",), required=True),
            _fm("X坐标", "x", aliases=("X",), data_type="float", transform="float(value)", required=True),
            _fm("Y坐标", "y", aliases=("Y",), data_type="float", transform="float(value)", required=True),
            _fm("孔口高程", "z_surface", aliases=("Z",), data_type="float", transform="float(value)", required=True),
            _fm("方位角", "azimuth", data_type="float", transform="float(value)", target_unit="°"),
            _fm("倾角", "dip", data_type="float", transform="float(value)", target_unit="°"),
            _fm("样品起深", "sample_from", data_type="float", transform="float(value)", required=True),
            _fm("样品止深", "sample_to", data_type="float", transform="float(value)", required=True),
            _fm("矿石类型", "ore_type", required=True),
            _fm("品位", "grade", aliases=("平均品位",), data_type="float", transform="float(value)"),
            _fm("密度", "density", data_type="float", transform="float(value)", target_unit="t/m³"),
        ),
        coordinate=_coord(description="支持矿区独立坐标与国家坐标的转换参数记录。"),
        attributes=(AttributeMappingRule("ore_type", "ore_type_std", {"矿石": "工业矿石", "低品位": "低品位矿石", "废石": "围岩", "夹石": "夹石"}),),
        output=_out("VTK", "三维建模软件", precision=3, layer_name="mine_orebody"),
        quality=_qc(("borehole_id", "x", "y", "z_surface", "sample_from", "sample_to", "ore_type"), ranges={"grade": [0, 100], "density": [0, 10], "azimuth": [0, 360], "dip": [-90, 90]}, max_missing=0.05),
    ))

    templates.append(_template(
        "水利水电工程地质模板", "水利水电工程", "地质剖面", ("CSV", "XLSX", "DXF", "GeoJSON"),
        "用于坝址、库区、地下厂房地质剖面、岩体分级及渗透参数。",
        tags=("水利水电", "坝址", "渗透", "岩体分级"),
        fields_=(
            _fm("对象编号", "object_id", aliases=("剖面编号", "单元编号"), required=True),
            _fm("工程部位", "engineering_part", required=True),
            _fm("岩性", "lithology", required=True),
            _fm("岩体等级", "rock_mass_class"),
            _fm("渗透系数", "permeability", aliases=("K",), data_type="float", transform="float(value)", target_unit="m/s"),
            _fm("吕荣值", "lugeon", aliases=("Lu",), data_type="float", transform="float(value)"),
            _fm("弹性模量", "elastic_modulus", aliases=("E",), data_type="float", transform="float(value)", target_unit="GPa"),
            _fm("泊松比", "poisson_ratio", aliases=("μ",), data_type="float", transform="float(value)"),
            _fm("断层编号", "fault_id"),
            _fm("地下水类型", "groundwater_type"),
        ),
        coordinate=_coord(),
        attributes=(AttributeMappingRule("rock_mass_class", "rock_mass_class_std", {"A": "A类", "B": "B类", "C": "C类", "D": "D类", "I": "I级", "II": "II级", "III": "III级", "IV": "IV级", "V": "V级"}),),
        output=_out("GPKG", "GIS平台", layer_name="hydropower_geology"),
        quality=_qc(("object_id", "engineering_part", "lithology"), ranges={"permeability": [0, None], "lugeon": [0, None], "poisson_ratio": [0, 0.5]}, max_missing=0.1),
    ))

    templates.append(_template(
        "城市地质三维建模模板", "城市地质调查", "钻孔数据", ("CSV", "XLSX", "JSON", "GeoJSON"),
        "用于城市第四系分层、基岩面、含水层及工程地质分区建模。",
        tags=("城市地质", "第四系", "基岩面", "工程地质分区"),
        fields_=(
            _fm("钻孔编号", "borehole_id", aliases=("孔号",), required=True),
            _fm("X坐标", "x", aliases=("X", "东坐标"), data_type="float", transform="float(value)", required=True),
            _fm("Y坐标", "y", aliases=("Y", "北坐标"), data_type="float", transform="float(value)", required=True),
            _fm("孔口高程", "z_surface", aliases=("高程",), data_type="float", transform="float(value)", required=True),
            _fm("层顶深度", "layer_top", data_type="float", transform="float(value)", default=0.0),
            _fm("层底深度", "layer_bottom", aliases=("分层深度",), data_type="float", transform="float(value)", required=True),
            _fm("地层名称", "strata_name", required=True),
            _fm("岩性名称", "lithology", aliases=("岩性",), required=True),
            _fm("地层时代", "strata_age"),
            _fm("标贯击数", "spt_n", aliases=("SPT_N",), data_type="float", transform="float(value)"),
            _fm("含水层", "aquifer"),
            _fm("工程地质分区", "engineering_zone"),
        ),
        coordinate=_coord(),
        attributes=(
            AttributeMappingRule("strata_age", "era_std", {"Qh": "全新世", "Qp": "更新世", "Q": "第四纪", "N": "新近纪", "E": "古近纪", "K": "白垩纪", "J": "侏罗纪", "T": "三叠纪"}),
            AttributeMappingRule("engineering_zone", "engineering_zone_std", {"I": "适宜区", "II": "较适宜区", "III": "一般适宜区", "IV": "限制区"}),
        ),
        output=_out("GeoJSON", "GIS平台", layer_name="urban_geology_boreholes"),
        quality=_qc(("borehole_id", "x", "y", "z_surface", "layer_bottom", "strata_name", "lithology"), ranges={"spt_n": [0, 200]}, max_missing=0.08),
    ))

    templates.append(_template(
        "区域地质图标准化模板", "区域地质调查", "地质图", ("SHP", "GPKG", "GeoJSON", "DXF", "CSV"),
        "用于地层面、断层线、产状点及区域地质图图层标准化。",
        tags=("区域地质", "地质图", "地层", "断层"),
        fields_=(
            _fm("要素编号", "feature_id", aliases=("ID", "编号"), required=True),
            _fm("要素类型", "feature_type", aliases=("类型",), required=True),
            _fm("地层代号", "strata_code", aliases=("地层代码",)),
            _fm("地层名称", "strata_name"),
            _fm("岩性", "lithology"),
            _fm("断层名称", "fault_name"),
            _fm("断层性质", "fault_type"),
            _fm("走向", "strike", data_type="float", transform="float(value)", target_unit="°"),
            _fm("倾向", "dip_direction", data_type="float", transform="float(value)", target_unit="°"),
            _fm("倾角", "dip_angle", data_type="float", transform="float(value)", target_unit="°"),
        ),
        coordinate=_coord(source_crs="WGS84", target_crs="CGCS2000", source_epsg="4326", description="实际投影转换由 GIS/GDAL 服务执行。"),
        attributes=(
            AttributeMappingRule("feature_type", "feature_type_std", {"面": "地质面", "地层面": "地质面", "线": "地质线", "断层": "断层线", "点": "地质点", "产状": "产状点"}),
            AttributeMappingRule("fault_type", "fault_type_std", {"正断层": "正断层", "逆断层": "逆断层", "走滑": "走滑断层", "平移断层": "走滑断层", "性质不明": "未定断层"}),
        ),
        output=_out("GPKG", "GIS平台", layer_name="regional_geology"),
        quality=_qc(("feature_id", "feature_type"), unique=("feature_id",), ranges={"strike": [0, 360], "dip_direction": [0, 360], "dip_angle": [0, 90]}, max_missing=0.1),
    ))

    templates.append(_template(
        "环境地质评价模板", "环境地质评价", "属性表", ("CSV", "XLSX", "GeoJSON", "SHP"),
        "用于地质灾害、地下水污染、地面沉降及生态地质评价指标标准化。",
        tags=("环境地质", "灾害", "污染", "评价"),
        fields_=(
            _fm("评价单元编号", "unit_id", aliases=("单元编号", "ID"), required=True),
            _fm("评价类型", "assessment_type", required=True),
            _fm("危险性", "hazard_level"),
            _fm("易发性", "susceptibility"),
            _fm("脆弱性", "vulnerability"),
            _fm("风险等级", "risk_level", required=True),
            _fm("污染物", "pollutant"),
            _fm("浓度", "concentration", data_type="float", transform="float(value)"),
            _fm("沉降速率", "subsidence_rate", data_type="float", transform="float(value)", target_unit="mm/a"),
            _fm("数据来源", "source"),
        ),
        coordinate=_coord(),
        attributes=(
            AttributeMappingRule("risk_level", "risk_level_std", {"低": "低风险", "较低": "较低风险", "中": "中风险", "较高": "较高风险", "高": "高风险"}),
            AttributeMappingRule("assessment_type", "assessment_type_std", {"滑坡": "地质灾害", "崩塌": "地质灾害", "泥石流": "地质灾害", "污染": "地下水污染", "沉降": "地面沉降"}),
        ),
        output=_out("GeoJSON", "GIS平台", layer_name="environmental_geology"),
        quality=_qc(("unit_id", "assessment_type", "risk_level"), unique=("unit_id",), max_missing=0.1),
    ))

    templates.append(_template(
        "点云标准化与分类模板", "通用工程", "点云数据", ("LAS", "LAZ", "PLY", "XYZ", "CSV"),
        "用于激光扫描或摄影测量点云的字段统一、分类编码和质量检查。",
        tags=("点云", "LAS", "分类", "抽稀"),
        fields_=(
            _fm("X", "x", aliases=("X坐标",), data_type="float", transform="float(value)", required=True, target_unit="m"),
            _fm("Y", "y", aliases=("Y坐标",), data_type="float", transform="float(value)", required=True, target_unit="m"),
            _fm("Z", "z", aliases=("Z坐标", "高程"), data_type="float", transform="float(value)", required=True, target_unit="m"),
            _fm("强度", "intensity", aliases=("Intensity",), data_type="integer", transform="int(value)"),
            _fm("分类", "classification", aliases=("Class", "类别"), data_type="integer", transform="int(value)"),
            _fm("回波序号", "return_number", aliases=("ReturnNumber",), data_type="integer", transform="int(value)"),
            _fm("回波总数", "number_of_returns", aliases=("NumberOfReturns",), data_type="integer", transform="int(value)"),
            _fm("红", "red", aliases=("R",), data_type="integer", transform="int(value)"),
            _fm("绿", "green", aliases=("G",), data_type="integer", transform="int(value)"),
            _fm("蓝", "blue", aliases=("B",), data_type="integer", transform="int(value)"),
        ),
        coordinate=_coord(),
        attributes=(AttributeMappingRule("classification", "classification_name", {"0": "未分类", "1": "未分类", "2": "地面", "3": "低植被", "4": "中植被", "5": "高植被", "6": "建筑物", "9": "水体", "17": "桥面"}),),
        output=_out("LAZ", "三维建模软件", precision=3, include_normals=False, layer_name="classified_point_cloud", options={"compression": True}),
        quality=_qc(("x", "y", "z"), ranges={"intensity": [0, 65535], "classification": [0, 255], "red": [0, 65535], "green": [0, 65535], "blue": [0, 65535]}, max_missing=0.02, topology=False, closure=False),
    ))

    templates.append(_template(
        "遥感影像与DEM成果模板", "区域地质调查", "遥感影像", ("TIFF", "GeoTIFF", "IMG", "ASC"),
        "用于遥感影像、DEM/DSM 栅格的坐标、像元、NoData 和输出压缩规则配置。",
        tags=("遥感", "DEM", "栅格", "GeoTIFF"),
        fields_=(
            _fm("影像编号", "raster_id", aliases=("ID",), required=True),
            _fm("分辨率", "resolution", aliases=("像元大小",), data_type="float", transform="float(value)", target_unit="m"),
            _fm("波段数", "band_count", data_type="integer", transform="int(value)"),
            _fm("无效值", "nodata", aliases=("NoData",), data_type="float", transform="float(value)"),
            _fm("采集日期", "acquisition_date", data_type="date"),
            _fm("传感器", "sensor"),
        ),
        coordinate=_coord(source_crs="WGS84", target_crs="CGCS2000", source_epsg="4326", description="栅格重投影由 GDAL 后端执行。"),
        output=_out("GeoTIFF", "GIS平台", precision=3, include_normals=False, include_properties=True, layer_name="remote_sensing", options={"compression": "LZW", "tiled": True, "build_overviews": True}),
        quality=_qc(("raster_id", "resolution"), ranges={"resolution": [0, None], "band_count": [1, 1000]}, max_missing=0.05, topology=False, closure=False),
    ))

    templates.append(_template(
        "地质剖面空间化模板", "通用工程", "地质剖面", ("DXF", "DWG", "CSV", "GeoJSON", "SVG"),
        "用于二维地质剖面控制点、里程高程、地层界线及断层线空间定位。",
        tags=("剖面", "空间化", "界线", "断层"),
        fields_=(
            _fm("剖面编号", "section_id", required=True),
            _fm("要素编号", "feature_id", required=True),
            _fm("要素类型", "feature_type", required=True),
            _fm("里程", "chainage", data_type="float", transform="float(value)", required=True, target_unit="m"),
            _fm("高程", "elevation", data_type="float", transform="float(value)", required=True, target_unit="m"),
            _fm("地层代号", "strata_code"),
            _fm("岩性", "lithology"),
            _fm("断层编号", "fault_id"),
        ),
        coordinate=CoordinateRule(source_crs="剖面局部坐标", target_crs="工程三维坐标", description="需结合剖面起点、方位角和控制点完成三维空间化。"),
        attributes=(AttributeMappingRule("feature_type", "feature_type_std", {"地层线": "地层界线", "界线": "地层界线", "断层": "断层线", "地表": "地形线", "钻孔": "钻孔投影"}),),
        output=_out("DXF", "三维建模软件", precision=3, layer_name="geological_section"),
        quality=_qc(("section_id", "feature_id", "feature_type", "chainage", "elevation"), unique=("feature_id",), max_missing=0.05, closure=False),
    ))

    templates.append(_template(
        "地质属性表标准化模板", "通用工程", "属性表", ("CSV", "XLSX", "JSON"),
        "将钻孔、剖面、点云、栅格、矢量、网格等来源的通用属性表统一为标准结构。",
        tags=("属性表", "字段映射", "标准化", "通用"),
        fields_=(
            _fm("编号", "id", aliases=("ID", "对象编号"), required=True),
            _fm("名称", "name", aliases=("对象名称",)),
            _fm("类型", "type", aliases=("对象类型",), required=True),
            _fm("描述", "description", aliases=("说明",)),
            _fm("来源", "source", aliases=("数据来源",)),
            _fm("单位", "unit"),
            _fm("备注", "remarks"),
        ),
        coordinate=None,
        attributes=(AttributeMappingRule("type", "type_std", {"钻孔": "borehole", "剖面": "section", "点云": "point_cloud", "栅格": "raster", "矢量": "vector", "网格": "mesh", "表格": "tabular", "模型": "model"}),),
        output=_out("CSV", "GIS平台", include_normals=False, layer_name="normalized_attributes"),
        quality=_qc(("id", "type"), unique=("id",), max_missing=0.15, topology=False, closure=False),
    ))

    templates.append(_template(
        "网格模型数值模拟输出模板", "通用工程", "网格模型", ("OBJ", "STL", "VTK", "VTU", "INP", "CSV"),
        "将地质网格模型转换为 FLAC3D、ABAQUS、ANSYS、OpenFOAM 等数值模拟成果。",
        tags=("网格", "数值模拟", "ABAQUS", "FLAC3D", "VTK"),
        fields_=(
            _fm("节点编号", "node_id", aliases=("NodeID",), data_type="integer", transform="int(value)", required=True),
            _fm("X", "x", aliases=("X坐标",), data_type="float", transform="float(value)", required=True),
            _fm("Y", "y", aliases=("Y坐标",), data_type="float", transform="float(value)", required=True),
            _fm("Z", "z", aliases=("Z坐标",), data_type="float", transform="float(value)", required=True),
            _fm("单元编号", "element_id", aliases=("ElementID",), data_type="integer", transform="int(value)"),
            _fm("单元类型", "element_type"),
            _fm("材料编号", "material_id", data_type="integer", transform="int(value)"),
            _fm("弹性模量", "elastic_modulus", aliases=("E",), data_type="float", transform="float(value)", target_unit="Pa"),
            _fm("泊松比", "poisson_ratio", aliases=("nu",), data_type="float", transform="float(value)"),
            _fm("密度", "density", aliases=("rho",), data_type="float", transform="float(value)", target_unit="kg/m³"),
            _fm("黏聚力", "cohesion", aliases=("c",), data_type="float", transform="float(value)", target_unit="Pa"),
            _fm("内摩擦角", "friction_angle", aliases=("phi",), data_type="float", transform="float(value)", target_unit="°"),
        ),
        coordinate=CoordinateRule(source_crs="工程独立坐标系", target_crs="模型局部坐标系", description="通过 offset 和 scale_factor 将大坐标平移至模拟局部坐标。"),
        output=_out("VTK", "数值模拟软件", precision=6, include_normals=False, layer_name="simulation_mesh", options={"cell_data": True, "material_groups": True}),
        quality=_qc(("node_id", "x", "y", "z"), unique=("node_id",), ranges={"poisson_ratio": [0, 0.5], "density": [0, None], "friction_angle": [0, 90]}, max_missing=0.03),
    ))

    templates.append(_template(
        "BIM地质模型交付模板", "通用工程", "网格模型", ("OBJ", "STL", "IFC", "RVT", "DWG", "CSV"),
        "用于地质体、钻孔和工程构筑物向 BIM 平台交付，保留分类、属性和坐标基点。",
        tags=("BIM", "IFC", "Revit", "交付"),
        fields_=(
            _fm("对象编号", "global_id", aliases=("GUID", "ID"), required=True),
            _fm("对象名称", "name", aliases=("名称",), required=True),
            _fm("对象类型", "object_type", aliases=("类型",), required=True),
            _fm("岩性", "lithology"),
            _fm("地层", "stratum"),
            _fm("材料", "material"),
            _fm("阶段", "phase"),
            _fm("来源模型", "source_model"),
        ),
        coordinate=CoordinateRule(source_crs="工程坐标系", target_crs="BIM共享坐标系", description="记录项目基点、测量点和真北旋转角。"),
        attributes=(AttributeMappingRule("object_type", "ifc_class", {"地质体": "IfcGeographicElement", "钻孔": "IfcProxy", "断层": "IfcGeographicElement", "隧道": "IfcTunnel", "构件": "IfcBuildingElementProxy"}, default_value="IfcProxy"),),
        output=_out("IFC", "BIM平台", precision=4, include_textures=True, layer_name="geology_bim", options={"ifc_schema": "IFC4.3", "georeference": True, "property_sets": True}),
        quality=_qc(("global_id", "name", "object_type"), unique=("global_id",), max_missing=0.05),
    ))

    templates.append(_template(
        "GIS地质成果输出模板", "通用工程", "通用数据", ("CSV", "GeoJSON", "SHP", "GPKG", "TIFF", "GeoTIFF"),
        "用于向 ArcGIS、QGIS 等 GIS 平台输出点、线、面、栅格及属性成果。",
        tags=("GIS", "GeoJSON", "GPKG", "ArcGIS", "QGIS"),
        fields_=(
            _fm("要素编号", "feature_id", aliases=("ID", "编号"), required=True),
            _fm("要素名称", "feature_name", aliases=("名称",)),
            _fm("图层", "layer", aliases=("图层名",), required=True),
            _fm("几何类型", "geometry_type", aliases=("类型",), required=True),
            _fm("数据来源", "source"),
            _fm("更新时间", "updated_at", data_type="date"),
        ),
        coordinate=_coord(),
        attributes=(AttributeMappingRule("geometry_type", "geometry_type_std", {"点": "Point", "线": "LineString", "面": "Polygon", "多点": "MultiPoint", "多线": "MultiLineString", "多面": "MultiPolygon"}),),
        output=_out("GPKG", "GIS平台", precision=6, include_normals=False, layer_name="geology_delivery", options={"spatial_index": True}),
        quality=_qc(("feature_id", "layer", "geometry_type"), unique=("feature_id",), max_missing=0.05),
    ))

    templates.append(_template(
        "三维建模软件交换模板", "通用工程", "网格模型", ("OBJ", "STL", "PLY", "FBX", "GLTF", "GLB"),
        "用于 Geomagic、Blender、3ds Max、MeshLab 等三维软件间交换地质表面与实体模型。",
        tags=("三维建模", "OBJ", "STL", "PLY", "glTF"),
        fields_=(
            _fm("对象编号", "object_id", aliases=("ID",), required=True),
            _fm("对象名称", "name", aliases=("名称",)),
            _fm("材质编号", "material_id"),
            _fm("岩性", "lithology"),
            _fm("分组", "group_name", aliases=("组",)),
        ),
        coordinate=CoordinateRule(source_crs="工程坐标系", target_crs="模型局部坐标系", description="建议保留原始大地坐标元数据，并将显示模型平移至局部原点。"),
        output=_out("OBJ", "三维建模软件", precision=6, include_normals=True, include_textures=True, layer_name="geology_mesh", options={"write_mtl": True, "triangulate": True}),
        quality=_qc(("object_id",), unique=("object_id",), max_missing=0.1),
    ))

    templates.append(_template(
        "地质模型全面质量检查模板", "通用工程", "通用数据", COMMON_INPUT_FORMATS,
        "交付前全面检查几何闭合、拓扑、属性完整性、一致性、坐标精度、重复数据和单位。",
        tags=("质量检查", "交付", "几何", "拓扑", "属性"),
        output=_out("JSON", "GIS平台", include_normals=False, layer_name="quality_report"),
        quality=QualityCheckRule(
            check_geometry_closure=True,
            check_topology=True,
            check_attribute_completeness=True,
            check_consistency=True,
            check_coordinate_precision=True,
            check_duplicates=True,
            check_units=True,
            closure_tolerance=0.001,
            outlier_sigma=3.0,
            min_face_quality=0.3,
            max_missing_ratio=0.1,
            description="适用于模型成果交付前的最终质量门禁。",
        ),
    ))

    return templates


_BUILTIN_TEMPLATES: dict[str, RuleTemplate] = {item.name: item for item in _build_builtin_templates()}
_USER_TEMPLATES: dict[str, RuleTemplate] = {}


# ============================================================================
# 模板校验、查询与持久化
# ============================================================================

def validate_template(template: RuleTemplate) -> TemplateValidationResult:
    errors: list[str] = []
    warnings: list[str] = []

    if not template.name.strip():
        errors.append("模板名称不能为空。")
    if template.project_type not in PROJECT_TYPES:
        errors.append(f"未知项目类型：{template.project_type}")
    if template.data_source not in DATA_SOURCES:
        errors.append(f"未知数据来源：{template.data_source}")
    if not template.input_formats:
        warnings.append("未声明输入格式。")

    targets: set[str] = set()
    for index, mapping in enumerate(template.field_mappings, start=1):
        if not mapping.source_field or not mapping.target_field:
            errors.append(f"第 {index} 条字段映射缺少源字段或目标字段。")
        if mapping.target_field in targets:
            warnings.append(f"目标字段重复：{mapping.target_field}")
        targets.add(mapping.target_field)
        if mapping.data_type not in {"string", "integer", "float", "boolean", "date"}:
            warnings.append(f"字段 {mapping.target_field} 的数据类型不常见：{mapping.data_type}")

    quality = template.quality_check_rule
    if quality:
        for required in quality.required_fields:
            if template.field_mappings and required not in targets:
                warnings.append(f"质检必填字段未出现在字段映射中：{required}")
        if not 0 <= quality.max_missing_ratio <= 1:
            errors.append("max_missing_ratio 必须位于 0～1。")

    if template.output_rule and template.output_rule.target_platform not in TARGET_PLATFORMS:
        errors.append(f"未知目标平台：{template.output_rule.target_platform}")

    return TemplateValidationResult(valid=not errors, errors=errors, warnings=warnings)


def _all_registry() -> dict[str, RuleTemplate]:
    merged = dict(_BUILTIN_TEMPLATES)
    merged.update(_USER_TEMPLATES)
    return merged


def list_templates(
    project_type: str | None = None,
    data_source: str | None = None,
    tag: str | None = None,
    include_disabled: bool = False,
) -> list[RuleTemplate]:
    result = list(_all_registry().values())
    if not include_disabled:
        result = [item for item in result if item.enabled]
    if project_type and project_type != "全部":
        result = [item for item in result if item.project_type in {project_type, "通用工程"}]
    if data_source and data_source != "全部":
        result = [item for item in result if item.data_source in {data_source, "通用数据"}]
    if tag:
        tag_lower = tag.lower()
        result = [item for item in result if any(tag_lower in value.lower() for value in item.tags)]
    return sorted((copy.deepcopy(item) for item in result), key=lambda item: (item.project_type, item.data_source, item.name))


def get_template(name: str) -> RuleTemplate | None:
    template = _all_registry().get(name)
    return copy.deepcopy(template) if template else None


def get_template_names() -> list[str]:
    return [item.name for item in list_templates()]


def save_template(template: RuleTemplate, storage_dir: str | Path | None = None) -> str:
    result = validate_template(template)
    if not result.valid:
        raise ValueError("；".join(result.errors))
    stored = copy.deepcopy(template)
    if not stored.created_at:
        stored.created_at = _now()
    stored.updated_at = _now()
    if stored.author == "系统预置":
        stored.author = "用户修改"
    _USER_TEMPLATES[stored.name] = stored
    if storage_dir is not None:
        safe_name = re.sub(r"[\/:*?\"<>|]+", "_", stored.name)
        _atomic_json(Path(storage_dir) / f"{safe_name}.json", stored.to_dict())
    return stored.name


def delete_template(name: str, storage_dir: str | Path | None = None) -> bool:
    if name in _BUILTIN_TEMPLATES:
        return False
    deleted = _USER_TEMPLATES.pop(name, None) is not None
    if deleted and storage_dir is not None:
        safe_name = re.sub(r"[\/:*?\"<>|]+", "_", name)
        (Path(storage_dir) / f"{safe_name}.json").unlink(missing_ok=True)
    return deleted


def duplicate_template(source_name: str, new_name: str, storage_dir: str | Path | None = None) -> RuleTemplate | None:
    source = get_template(source_name)
    if not source:
        return None
    source.name = new_name
    source.version = f"{source.version}+copy"
    source.author = "用户复制"
    source.created_at = _now()
    source.updated_at = _now()
    save_template(source, storage_dir=storage_dir)
    return source


def export_template(name: str, output_path: str | Path) -> str | None:
    template = get_template(name)
    if not template:
        return None
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    _atomic_json(output, template.to_dict())
    return str(output.resolve())


def export_template_library(output_dir: str | Path) -> list[str]:
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    exported: list[str] = []
    index_rows: list[dict[str, Any]] = []
    for template in list_templates(include_disabled=True):
        safe_name = re.sub(r"[\\/:*?\"<>|]+", "_", template.name)
        path = output / f"{safe_name}.json"
        _atomic_json(path, template.to_dict())
        exported.append(str(path.resolve()))
        index_rows.append({
            "name": template.name,
            "version": template.version,
            "project_type": template.project_type,
            "data_source": template.data_source,
            "input_formats": template.input_formats,
            "file": path.name,
            "sha256": _sha256(path),
            "fingerprint": template_fingerprint(template),
            "validation": asdict(validate_template(template)),
        })
    index_path = output / "template_index.json"
    _atomic_json(index_path, {
        "schema_version": SCHEMA_VERSION,
        "exported_at": _now(),
        "template_count": len(index_rows),
        "templates": index_rows,
    })
    exported.append(str(index_path.resolve()))
    return exported


def import_template(file_path: str | Path, storage_dir: str | Path | None = None) -> str | None:
    path = Path(file_path)
    if not path.exists():
        return None
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    template = RuleTemplate.from_dict(data)
    result = validate_template(template)
    if not result.valid:
        raise ValueError("；".join(result.errors))
    if not template.created_at:
        template.created_at = _now()
    template.updated_at = _now()
    if template.author == "系统预置":
        template.author = "外部导入"
    return save_template(template, storage_dir=storage_dir)


def load_template_directory(directory: str | Path, storage_dir: str | Path | None = None) -> dict[str, list[str]]:
    root = Path(directory)
    imported: list[str] = []
    errors: list[str] = []
    if not root.exists():
        return {"imported": imported, "errors": [f"目录不存在：{root}"]}
    for path in sorted(root.glob("*.json")):
        if path.name == "template_index.json":
            continue
        try:
            name = import_template(path, storage_dir=storage_dir)
            if name:
                imported.append(name)
        except Exception as exc:
            errors.append(f"{path.name}: {exc}")
    return {"imported": imported, "errors": errors}


# ============================================================================
# 安全转换与模板应用
# ============================================================================

_ALLOWED_BINARY_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.Mod: operator.mod,
    ast.Pow: operator.pow,
}
_ALLOWED_UNARY_OPERATORS = {ast.UAdd: operator.pos, ast.USub: operator.neg}
_ALLOWED_FUNCTIONS = {
    "float": float,
    "int": int,
    "str": str,
    "round": round,
    "abs": abs,
    "min": min,
    "max": max,
}
_ALLOWED_STRING_METHODS = {"strip", "upper", "lower", "title", "replace"}


def _safe_eval_transform(expression: str, value: Any) -> Any:
    tree = ast.parse(expression, mode="eval")

    def evaluate(node: ast.AST) -> Any:
        if isinstance(node, ast.Expression):
            return evaluate(node.body)
        if isinstance(node, ast.Name):
            if node.id == "value":
                return value
            if node.id in _ALLOWED_FUNCTIONS:
                return _ALLOWED_FUNCTIONS[node.id]
            raise ValueError(f"不允许的名称：{node.id}")
        if isinstance(node, ast.Constant):
            return node.value
        if isinstance(node, ast.BinOp) and type(node.op) in _ALLOWED_BINARY_OPERATORS:
            return _ALLOWED_BINARY_OPERATORS[type(node.op)](evaluate(node.left), evaluate(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in _ALLOWED_UNARY_OPERATORS:
            return _ALLOWED_UNARY_OPERATORS[type(node.op)](evaluate(node.operand))
        if isinstance(node, ast.Call):
            args = [evaluate(arg) for arg in node.args]
            if isinstance(node.func, ast.Name) and node.func.id in _ALLOWED_FUNCTIONS:
                return _ALLOWED_FUNCTIONS[node.func.id](*args)
            if isinstance(node.func, ast.Attribute):
                owner = evaluate(node.func.value)
                if isinstance(owner, str) and node.func.attr in _ALLOWED_STRING_METHODS:
                    return getattr(owner, node.func.attr)(*args)
            raise ValueError("不允许的函数调用。")
        raise ValueError(f"不允许的表达式节点：{type(node).__name__}")

    return evaluate(tree)


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return isinstance(value, str) and value.strip().lower() in MISSING_MARKERS


def _find_source_value(row: dict[str, Any], mapping: FieldMappingRule) -> tuple[Any, str | None]:
    candidates = [mapping.source_field, *mapping.aliases]
    direct = {str(key): key for key in row}
    normalized = {str(key).strip().lower(): key for key in row}
    for candidate in candidates:
        if candidate in direct:
            return row[direct[candidate]], str(direct[candidate])
        matched_key = normalized.get(candidate.strip().lower())
        if matched_key is not None:
            return row[matched_key], str(matched_key)
    return None, None


def _normalize_map_value(value: Any, rule: AttributeMappingRule) -> str:
    text = "" if value is None else str(value)
    if rule.strip_value:
        text = text.strip()
    if rule.case_sensitive:
        return rule.value_map.get(text, rule.default_value or text)
    lookup = {str(key).lower(): mapped for key, mapped in rule.value_map.items()}
    return lookup.get(text.lower(), rule.default_value or text)


def apply_template_to_data(template_name: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    template = get_template(template_name)
    if not template:
        return {"rows": rows, "fields_mapped": [], "errors": ["模板不存在。"], "warnings": [], "quality": {}}

    mapped_rows: list[dict[str, Any]] = []
    errors: list[str] = []
    warnings: list[str] = []

    for row_index, row in enumerate(rows, start=1):
        mapped: dict[str, Any] = {}
        for mapping in template.field_mappings:
            raw, source_key = _find_source_value(row, mapping)
            if _is_missing(raw):
                mapped[mapping.target_field] = mapping.default_value
                if mapping.required and mapping.default_value is None:
                    errors.append(f"第 {row_index} 行缺少必填字段“{mapping.source_field}”。")
                continue
            try:
                mapped_value = _safe_eval_transform(mapping.transform, raw) if mapping.transform else raw
            except Exception as exc:
                mapped_value = raw
                warnings.append(
                    f"第 {row_index} 行字段“{source_key or mapping.source_field}”转换失败，已保留原值：{exc}"
                )
            mapped[mapping.target_field] = mapped_value

        for attr_rule in template.attribute_mappings:
            source_value = mapped.get(attr_rule.source_field)
            if source_value is None:
                source_value = row.get(attr_rule.source_field)
            mapped[attr_rule.target_field] = _normalize_map_value(source_value, attr_rule)

        coordinate = template.coordinate_rule
        if coordinate:
            for target, offset in (("x", coordinate.offset_x), ("y", coordinate.offset_y), ("z", coordinate.offset_z)):
                value = mapped.get(target)
                if isinstance(value, (int, float)):
                    mapped[target] = round(value * coordinate.scale_factor + offset, coordinate.precision)

        mapped_rows.append(mapped)

    quality = _run_tabular_quality_checks(mapped_rows, template.quality_check_rule)
    errors.extend(quality["errors"])
    warnings.extend(quality["warnings"])

    return {
        "template": template.name,
        "version": template.version,
        "rows": mapped_rows,
        "fields_mapped": [mapping.target_field for mapping in template.field_mappings],
        "errors": errors,
        "warnings": warnings,
        "quality": quality,
    }


def _run_tabular_quality_checks(rows: list[dict[str, Any]], rule: QualityCheckRule | None) -> dict[str, Any]:
    result: dict[str, Any] = {
        "row_count": len(rows),
        "errors": [],
        "warnings": [],
        "missing_ratio": {},
        "duplicate_count": 0,
        "range_violation_count": 0,
        "allowed_value_violation_count": 0,
    }
    if not rule or not rows:
        return result

    fields_to_check = sorted({key for row in rows for key in row} | set(rule.required_fields))
    for name in fields_to_check:
        missing = sum(1 for row in rows if _is_missing(row.get(name)))
        ratio = missing / len(rows)
        result["missing_ratio"][name] = ratio
        if name in rule.required_fields and missing:
            result["errors"].append(f"必填字段 {name} 有 {missing} 条缺失记录。")
        elif ratio > rule.max_missing_ratio:
            result["warnings"].append(f"字段 {name} 缺失率 {ratio:.1%} 超过阈值 {rule.max_missing_ratio:.1%}。")

    for unique_field in rule.unique_fields:
        field_names = [name.strip() for name in unique_field.split("+") if name.strip()]
        seen: set[tuple[str, ...]] = set()
        duplicates = 0
        for row in rows:
            values = [row.get(name) for name in field_names]
            if not field_names or any(_is_missing(value) for value in values):
                continue
            marker = tuple(str(value) for value in values)
            if marker in seen:
                duplicates += 1
            seen.add(marker)
        if duplicates:
            result["duplicate_count"] += duplicates
            label = " + ".join(field_names)
            result["errors"].append(f"唯一键 {label} 存在 {duplicates} 条重复记录。")

    for field_name, bounds in rule.numeric_ranges.items():
        lower = bounds[0] if bounds else None
        upper = bounds[1] if len(bounds) > 1 else None
        for row_index, row in enumerate(rows, start=1):
            value = row.get(field_name)
            if _is_missing(value):
                continue
            try:
                number = float(value)
            except (TypeError, ValueError):
                result["range_violation_count"] += 1
                result["warnings"].append(f"第 {row_index} 行字段 {field_name} 不是数值：{value}")
                continue
            if (lower is not None and number < lower) or (upper is not None and number > upper):
                result["range_violation_count"] += 1
                result["warnings"].append(f"第 {row_index} 行字段 {field_name}={number} 超出范围 [{lower}, {upper}]。")

    for field_name, allowed in rule.allowed_values.items():
        allowed_set = {str(item) for item in allowed}
        for row_index, row in enumerate(rows, start=1):
            value = row.get(field_name)
            if _is_missing(value):
                continue
            if str(value) not in allowed_set:
                result["allowed_value_violation_count"] += 1
                result["warnings"].append(f"第 {row_index} 行字段 {field_name}={value} 不在允许值集合中。")

    return result


def _read_tabular_file(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("读取 XLSX 需要安装 openpyxl。") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(iterator, ())]
        rows: list[dict[str, Any]] = []
        for values in iterator:
            rows.append({headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]})
        workbook.close()
        return rows
    if suffix in {".json", ".geojson"}:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(data, list):
            return [item for item in data if isinstance(item, dict)]
        if isinstance(data, dict) and isinstance(data.get("features"), list):
            rows: list[dict[str, Any]] = []
            for feature in data["features"]:
                if not isinstance(feature, dict):
                    continue
                row = dict(feature.get("properties") or {})
                geometry = feature.get("geometry")
                if geometry:
                    row["geometry"] = geometry
                rows.append(row)
            return rows
        if isinstance(data, dict):
            return [data]
    raise ValueError(f"当前规则执行器支持 CSV/XLSX/JSON/GeoJSON，暂不直接解析：{path.suffix}")


def apply_template_to_file(
    template_name: str,
    input_path: str | Path,
    output_dir: str | Path,
) -> dict[str, Any]:
    input_file = Path(input_path)
    rows = _read_tabular_file(input_file)
    result = apply_template_to_data(template_name, rows)
    template = get_template(template_name)
    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    safe_template = re.sub(r"[\\/:*?\"<>|]+", "_", template_name)
    output_file = output_root / f"{input_file.stem}_{safe_template}_mapped.json"
    result.update({
        "schema_version": SCHEMA_VERSION,
        "application_id": f"TPL-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8].upper()}",
        "applied_at": _now(),
        "input_file": str(input_file.resolve()),
        "input_sha256": _sha256(input_file),
        "template_fingerprint": template_fingerprint(template) if template else "",
        "output_file": str(output_file.resolve()),
    })
    _atomic_json(output_file, result)
    result["output_sha256"] = _sha256(output_file)
    return result


# ============================================================================
# 文件识别与模板推荐
# ============================================================================

_EXTENSION_FORMATS = {
    ".csv": "CSV", ".xlsx": "XLSX", ".xls": "XLSX", ".json": "JSON", ".geojson": "GeoJSON",
    ".shp": "SHP", ".gpkg": "GPKG", ".dxf": "DXF", ".dwg": "DWG", ".obj": "OBJ",
    ".stl": "STL", ".ply": "PLY", ".las": "LAS", ".laz": "LAZ", ".tif": "GeoTIFF",
    ".tiff": "GeoTIFF", ".vtk": "VTK", ".vtu": "VTU", ".inp": "INP", ".ifc": "IFC",
    ".txt": "TXT", ".xyz": "XYZ",
}

_SOURCE_HINTS = {
    "钻孔数据": ("钻孔", "孔号", "孔深", "borehole", "hole_id", "rqd"),
    "地质图": ("地层代号", "断层", "产状", "strata", "fault", "polygon"),
    "物探数据": ("tsp", "gpr", "tem", "地质雷达", "瞬变电磁", "预报", "综合结论"),
    "遥感影像": ("dem", "dsm", "影像", "栅格", "像元", "nodata", "geotiff"),
    "点云数据": ("point cloud", "点云", "intensity", "classification", "las", "laz"),
    "地质剖面": ("剖面", "section", "坡高", "坡度", "里程", "高程"),
    "属性表": ("属性", "编号", "名称", "类型", "description"),
    "网格模型": ("node_id", "element_id", "节点编号", "单元编号", "mesh", "obj", "vtk"),
}

_PROJECT_HINTS = {
    "隧道工程": ("隧道", "围岩", "掌子面", "tsp", "gpr", "tem", "里程"),
    "边坡工程": ("边坡", "坡高", "坡度", "稳定性", "节理"),
    "地基与基础工程": ("地基", "基础", "标贯", "承载力", "土层"),
    "矿山工程": ("矿山", "矿体", "品位", "矿石", "样品起深"),
    "水利水电工程": ("水利", "水电", "坝址", "吕荣", "渗透系数"),
    "城市地质调查": ("城市地质", "第四系", "工程地质分区", "基岩面"),
    "区域地质调查": ("区域地质", "地质图", "地层代号", "断层性质"),
    "环境地质评价": ("环境地质", "污染", "沉降", "风险等级", "危险性"),
}


def _file_preview(path: Path) -> tuple[list[str], str]:
    headers: list[str] = []
    text = ""
    try:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                headers = next(reader, [])
                lines = [",".join(headers)]
                for _ in range(3):
                    row = next(reader, None)
                    if row is None:
                        break
                    lines.append(",".join(row))
                text = " ".join(lines)
        elif suffix in {".json", ".geojson"}:
            raw = path.read_text(encoding="utf-8-sig")[:20000]
            data = json.loads(raw) if len(raw) < 20000 else None
            if isinstance(data, dict):
                headers = list(data.keys())
                if isinstance(data.get("features"), list) and data["features"]:
                    props = data["features"][0].get("properties", {})
                    if isinstance(props, dict):
                        headers.extend(str(key) for key in props)
            text = raw
        elif suffix in {".txt", ".obj", ".vtk", ".inp", ".xyz"}:
            text = path.read_text(encoding="utf-8-sig", errors="ignore")[:10000]
    except Exception:
        pass
    return headers, text


def analyze_input_files(file_paths: list[str]) -> list[dict[str, Any]]:
    analyses: list[dict[str, Any]] = []
    for file_path in file_paths:
        path = Path(file_path)
        fmt = _EXTENSION_FORMATS.get(path.suffix.lower(), path.suffix.lstrip(".").upper() or "UNKNOWN")
        headers, preview = _file_preview(path) if path.exists() else ([], "")
        corpus = f"{path.name} {' '.join(headers)} {preview}".lower()
        source_scores = {
            source: sum(1 for hint in hints if hint.lower() in corpus)
            for source, hints in _SOURCE_HINTS.items()
        }
        project_scores = {
            project: sum(1 for hint in hints if hint.lower() in corpus)
            for project, hints in _PROJECT_HINTS.items()
        }
        inferred_source = max(source_scores, key=source_scores.get) if source_scores and max(source_scores.values()) > 0 else "通用数据"
        inferred_project = max(project_scores, key=project_scores.get) if project_scores and max(project_scores.values()) > 0 else "通用工程"
        analyses.append({
            "path": str(path),
            "name": path.name,
            "exists": path.exists(),
            "size": path.stat().st_size if path.exists() else 0,
            "sha256": _sha256(path) if path.exists() and path.is_file() else "",
            "format": fmt,
            "headers": headers,
            "inferred_data_source": inferred_source,
            "inferred_project_type": inferred_project,
        })
    return analyses


def audit_template_library() -> dict[str, Any]:
    templates = list_templates(include_disabled=True)
    invalid: list[dict[str, Any]] = []
    fingerprints: dict[str, list[str]] = {}
    for template in templates:
        validation = validate_template(template)
        if not validation.valid:
            invalid.append({"name": template.name, "errors": validation.errors, "warnings": validation.warnings})
        fingerprint = template_fingerprint(template)
        fingerprints.setdefault(fingerprint, []).append(template.name)
    duplicates = [
        {"fingerprint": fingerprint, "templates": names}
        for fingerprint, names in fingerprints.items() if len(names) > 1
    ]
    return {
        "valid": not invalid,
        "template_count": len(templates),
        "invalid_templates": invalid,
        "duplicate_content": duplicates,
        "built_in_count": len(_BUILTIN_TEMPLATES),
        "user_count": len(_USER_TEMPLATES),
    }


def build_template_report(
    file_paths: list[str],
    project_type: str = "全部",
    data_source: str = "全部",
) -> dict[str, Any]:
    file_analysis = analyze_input_files(file_paths)
    inferred_projects = [item["inferred_project_type"] for item in file_analysis if item["inferred_project_type"] != "通用工程"]
    inferred_sources = [item["inferred_data_source"] for item in file_analysis if item["inferred_data_source"] != "通用数据"]
    effective_project = project_type if project_type != "全部" else (_most_common(inferred_projects) or "全部")
    effective_source = data_source if data_source != "全部" else (_most_common(inferred_sources) or "全部")

    candidates = list_templates(
        project_type=None if effective_project == "全部" else effective_project,
        data_source=None if effective_source == "全部" else effective_source,
    )
    input_formats = {item["format"] for item in file_analysis}
    input_headers = {str(header).strip().lower() for item in file_analysis for header in item.get("headers", [])}

    matching: list[dict[str, Any]] = []
    for template in candidates:
        score = 0
        reasons: list[str] = []
        if effective_project != "全部":
            if template.project_type == effective_project:
                score += 35
                reasons.append(f"项目类型“{effective_project}”完全匹配")
            elif template.project_type == "通用工程":
                score += 12
                reasons.append("通用工程模板可兼容")
        if effective_source != "全部":
            if template.data_source == effective_source:
                score += 25
                reasons.append(f"数据来源“{effective_source}”完全匹配")
            elif template.data_source == "通用数据":
                score += 8
                reasons.append("通用数据模板可兼容")

        format_matches = sorted(input_formats.intersection(set(template.input_formats)))
        if format_matches:
            score += min(15, 5 + 5 * len(format_matches))
            reasons.append(f"支持输入格式：{', '.join(format_matches)}")

        if input_headers and template.field_mappings:
            matched_fields = 0
            for mapping in template.field_mappings:
                names = {mapping.source_field.lower(), *(alias.lower() for alias in mapping.aliases)}
                if input_headers.intersection(names):
                    matched_fields += 1
            coverage = matched_fields / len(template.field_mappings)
            score += round(20 * coverage)
            if matched_fields:
                reasons.append(f"字段命中 {matched_fields}/{len(template.field_mappings)}")
        else:
            coverage = 0.0

        rule_categories = sum((
            bool(template.field_mappings), bool(template.coordinate_rule), bool(template.attribute_mappings),
            bool(template.output_rule), bool(template.quality_check_rule),
        ))
        score += rule_categories
        score = min(100, score)

        matching.append({
            "name": template.name,
            "version": template.version,
            "project_type": template.project_type,
            "data_source": template.data_source,
            "input_formats": template.input_formats,
            "description": template.description,
            "score": score,
            "reasons": reasons,
            "rule_count": rule_categories,
            "field_coverage": coverage,
            "tag_count": len(template.tags),
        })

    matching.sort(key=lambda item: (item["score"], item["field_coverage"], item["version"]), reverse=True)
    all_templates = [_template_summary(item) for item in list_templates()]

    return {
        "schema_version": SCHEMA_VERSION,
        "run_id": f"MATCH-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8].upper()}",
        "checked_at": _now(),
        "project_type": project_type,
        "data_source": data_source,
        "effective_project_type": effective_project,
        "effective_data_source": effective_source,
        "files": file_analysis,
        "matching_templates": matching,
        "all_templates": all_templates,
        "library_audit": audit_template_library(),
        "summary": {
            "total_templates": len(all_templates),
            "matching_count": len(matching),
            "best_match": matching[0]["name"] if matching else None,
            "best_score": matching[0]["score"] if matching else 0,
            "project_types_available": sorted({item.project_type for item in list_templates()}),
            "data_sources_available": sorted({item.data_source for item in list_templates()}),
        },
    }


def _most_common(values: list[str]) -> str | None:
    if not values:
        return None
    return max(set(values), key=values.count)


def _template_summary(template: RuleTemplate) -> dict[str, Any]:
    return {
        "name": template.name,
        "version": template.version,
        "project_type": template.project_type,
        "data_source": template.data_source,
        "input_formats": template.input_formats,
        "rule_count": sum((
            bool(template.field_mappings), bool(template.coordinate_rule), bool(template.attribute_mappings),
            bool(template.output_rule), bool(template.quality_check_rule),
        )),
        "field_count": len(template.field_mappings),
        "attribute_rule_count": len(template.attribute_mappings),
        "author": template.author,
        "tags": template.tags,
    }


def get_example_files(source_dir: str | Path, project_type: str = "全部", data_source: str = "全部") -> list[str]:
    root = Path(source_dir) / "rule_template_examples"
    manifest = root / "example_manifest.json"
    if not manifest.exists():
        return []
    data = json.loads(manifest.read_text(encoding="utf-8-sig"))
    selected: list[str] = []
    for item in data:
        if item.get("kind", "data") != "data":
            continue
        if project_type not in {"全部", item.get("project_type")}:
            continue
        if data_source not in {"全部", item.get("data_source")}:
            continue
        path = root / item["file"]
        if path.exists():
            selected.append(str(path.resolve()))
    if selected:
        return selected
    return [str((root / item["file"]).resolve()) for item in data if item.get("kind", "data") == "data" and (root / item["file"]).exists()]


__all__ = [
    "SCHEMA_VERSION", "PROJECT_TYPES", "DATA_SOURCES", "TARGET_PLATFORMS", "COMMON_INPUT_FORMATS",
    "RuleTemplate", "FieldMappingRule", "CoordinateRule", "AttributeMappingRule", "OutputRule",
    "QualityCheckRule", "TemplateValidationResult", "validate_template",
    "list_templates", "get_template", "get_template_names", "save_template", "delete_template",
    "duplicate_template", "export_template", "export_template_library", "import_template",
    "load_template_directory", "apply_template_to_data", "apply_template_to_file",
    "analyze_input_files", "build_template_report", "get_example_files",
]
